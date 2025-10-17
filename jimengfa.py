# --- START OF FILE jimeng003.py (with model selection) ---

# Render 无 GUI 环境：将 tkinter 设为可选导入
try:
    import tkinter as tk
    from tkinter import ttk, scrolledtext, messagebox, filedialog
    TK_AVAILABLE = True
except Exception:
    # 在无 GUI 的服务器环境（如 Render）下，提供最小占位以避免导入错误
    TK_AVAILABLE = False
    from types import SimpleNamespace
    class _DummyTk: pass
    class _DummyMsg:
        @staticmethod
        def askyesno(*args, **kwargs): return False
        @staticmethod
        def showerror(*args, **kwargs): pass
        @staticmethod
        def showinfo(*args, **kwargs): pass
        @staticmethod
        def showwarning(*args, **kwargs): pass
    tk = SimpleNamespace(Tk=_DummyTk, BOTH=None, END=None, NORMAL=None, DISABLED=None)
    ttk = SimpleNamespace(Frame=object, Label=object, Button=object, LabelFrame=object, Entry=object,
                          Notebook=object, Combobox=object, Spinbox=object, Checkbutton=object,
                          Progressbar=object, Style=object)
    scrolledtext = SimpleNamespace(ScrolledText=object)
    filedialog = SimpleNamespace(askopenfilename=lambda **k: "", askdirectory=lambda **k: "",
                                 asksaveasfilename=lambda **k: "")
    messagebox = _DummyMsg
import json
import threading
import queue
import time
import hashlib
import requests
import uuid
import random
import zlib
import os
from urllib.parse import urlencode
import hmac
from datetime import datetime
from urllib.parse import urlparse
from typing import Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv


# ==============================================================================
# API 核心逻辑 (稳定版)
# ==============================================================================
class CapCutAPI:
    @staticmethod
    def get_web_id(cookies: dict) -> Tuple[Optional[str], Optional[str]]:
        """步骤0: 自动获取最新的 web_id (用作 did)"""
        url = 'https://mcs-normal-sg.capcutapi.com/v1/user/webid'
        payload = {
            "app_id": 611736, "url": "https://dreamina.capcut.com/",
            "user_agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            "referer": "", "user_unique_id": ""
        }
        headers = {'content-type': 'application/json; charset=UTF-8'}
        try:
            response = requests.post(url, headers=headers, json=payload, cookies=cookies, timeout=10)
            response.raise_for_status()
            data = response.json()
            if data.get("e") == 0 and data.get("web_id"):
                return data["web_id"], None
            else:
                return None, f"获取web_id失败: {data.get('message', '响应格式不正确')}"
        except requests.exceptions.RequestException as e:
            return None, f"请求web_id失败: {e}"

    def __init__(self, cookies_str, did):
        self.cookies = self._parse_cookies(cookies_str or "")
        
        _did = (did or "").strip()
        if not _did:
            print("[INFO] DID为空，尝试自动获取 web_id...")
            new_did, err = self.get_web_id(self.cookies)
            if err:
                print(f"[WARNING] 自动获取web_id失败: {err}")
                self.did = str(random.randint(10**18, 10**19 - 1)) # 如果获取失败，生成随机DID
                print(f"[INFO] 已生成随机DID: {self.did}")
            else:
                print(f"[SUCCESS] 成功获取 web_id: {new_did}")
                self.did = new_did
        else:
            self.did = _did

        self.session = requests.Session()
        self.common_url_params = {
            'aid': '513641', 'device_platform': 'web', 'region': 'SG', 'da_version': '3.2.8',
            'web_component_open_flag': '0', 'web_version': '6.6.0', 'aigc_features': 'app_lip_sync'
        }
        self.pf = "7"
        self.appvr = "5.8.0"
        self.user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'

    def _parse_cookies(self, cookies_str):
        cookies = {}
        try:
            text = cookies_str.strip()
            if text and ('=' not in text) and (';' not in text):
                cookies['sessionid'] = text
                cookies['sessionid_ss'] = text
                cookies['sid_tt'] = text
                return cookies
            for item in cookies_str.split(';'):
                item = item.strip()
                if not item: continue
                key, value = item.split('=', 1)
                cookies[key.strip()] = value.strip()
        except Exception as e:
            print(f"解析Cookie时出错: {e}")
        return cookies

    def _generate_sign(self, url_params, device_time, request_body_str, tdid_for_sign):
        sorted_params_str = urlencode(sorted(url_params.items())) if url_params else ""
        salt1 = "9e2c"; salt2 = "11ac"
        body_md5 = hashlib.md5(request_body_str.encode('utf-8')).hexdigest()
        string_to_sign = f"{salt1}|{sorted_params_str}|{self.pf}|{self.appvr}|{device_time}|{tdid_for_sign}|{body_md5}|{salt2}"
        return hashlib.md5(string_to_sign.encode('utf-8')).hexdigest()

    def _send_request(self, url, url_params, payload_str, tdid_for_sign, timeout=60):
        current_timestamp_sec = str(int(time.time()))
        signature = self._generate_sign(url_params, current_timestamp_sec, payload_str, tdid_for_sign)
        headers = {
            'accept': 'application/json, text/plain, */*', 'content-type': 'application/json',
            'device-time': current_timestamp_sec, 'did': self.did, 'pf': self.pf, 'appvr': self.appvr,
            'appid': '513641',
            'sign': signature, 'sign-ver': '1', 'referer': 'https://dreamina.capcut.com/',
            'user-agent': self.user_agent,
        }
        try:
            response = self.session.post(url, params=url_params, headers=headers, cookies=self.cookies, data=payload_str.encode('utf-8'), timeout=timeout)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            error_details = ""
            if e.response:
                try: error_details = e.response.json()
                except json.JSONDecodeError: error_details = e.response.text
            return {"error": f"网络请求失败: {e}", "details": error_details}
            
    def get_user_credit(self):
        api_url = "https://commerce-api-sg.capcut.com/commerce/v1/benefits/user_credit"
        payload_str = '{}'
        return self._send_request(api_url, None, payload_str, tdid_for_sign=self.did, timeout=10)

    def credit_receive(self, time_zone="Asia/Shanghai"):
        """领取每日积分（签到）"""
        api_url = "https://commerce-api-sg.capcut.com/commerce/v1/benefits/credit_receive"
        payload = {"time_zone": time_zone}
        payload_str = json.dumps(payload)
        return self._send_request(api_url, None, payload_str, tdid_for_sign=self.did, timeout=10)

    def get_common_config(self):
        """获取通用配置信息（包括所有可用模型列表）"""
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/get_common_config"
        # 使用更新的 URL 参数
        url_params = {
            'needCache': 'true',
            'needRefresh': 'false',
            'aid': '513641',
            'web_version': '7.5.0',
            'da_version': '3.3.2',
            'aigc_features': 'app_lip_sync'
        }
        payload_str = '{}'
        return self._send_request(api_url, url_params, payload_str, tdid_for_sign="web", timeout=10)

    def get_user_info_region(self, hashed_id="", temporary_id="", verify_fp=""):
        """获取用户信息和地区验证"""
        api_url = "https://login-row.www.capcut.com/passport/web/region/"
        
        url_params = {
            'aid': '513641',
            'account_sdk_source': 'web',
            'sdk_version': '2.1.10-tiktok',
            'language': 'en',
            'verifyFp': verify_fp or f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:8]}",
            'mix_mode': '1'
        }
        
        # POST 数据使用表单格式
        post_data = f"type=2&hashed_id={hashed_id}&temporary_id={temporary_id}"
        
        # 这个接口不需要签名，使用普通的 POST 请求
        headers = {
            'accept': 'application/json, text/plain, */*',
            'content-type': 'application/x-www-form-urlencoded',
            'did': self.did,
            'appid': '513641',
            'referer': 'https://dreamina.capcut.com/',
            'user-agent': self.user_agent,
        }
        
        try:
            response = self.session.post(api_url, params=url_params, headers=headers, 
                                        cookies=self.cookies, data=post_data, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            error_details = ""
            if e.response:
                try: error_details = e.response.json()
                except json.JSONDecodeError: error_details = e.response.text
            return {"error": f"获取用户信息失败: {e}", "details": error_details}

    def _encode_login_data(self, text):
        """
        编码登录数据（模拟前端加密逻辑）
        算法：字符串 → UTF-8字节 → 每个字节XOR 5 → 转十六进制
        对应前端代码中的 Q 函数
        """
        if not text:
            return ""
        
        # 1. 将字符串转换为 UTF-8 字节数组
        utf8_bytes = text.encode('utf-8')
        
        # 2. 对每个字节与 5 进行异或操作，然后转换为十六进制
        hex_result = []
        for byte in utf8_bytes:
            xor_byte = byte ^ 5  # 与 5 异或
            hex_result.append(format(xor_byte, 'x'))  # 转十六进制（不补0）
        
        return ''.join(hex_result)
    
    def login_with_email(self, email, password):
        """通过邮箱密码登录获取 session id（完整流程）"""
        # 生成验证指纹
        verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
        
        try:
            # 步骤1: 地区验证
            region_url = "https://login-row.www.capcut.com/passport/web/region/"
            region_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp,
                'mix_mode': '1'
            }
            
            # 编码邮箱和临时ID
            encoded_email = self._encode_login_data(email)
            temp_id = self._encode_login_data(f"temp_{email}")
            hashed_id = hashlib.sha256(email.encode()).hexdigest()
            
            region_data = f"type=2&hashed_id={hashed_id}&temporary_id={temp_id}"
            
            region_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': self.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/',
                'user-agent': self.user_agent,
            }
            
            region_response = self.session.post(
                region_url,
                params=region_params,
                headers=region_headers,
                data=region_data,
                timeout=10
            )
            
            # 步骤2: 邮箱登录
            login_url = "https://dreamina.capcut.com/passport/web/email/login/"
            login_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            # 对邮箱和密码进行加密
            encoded_email = self._encode_login_data(email)
            encoded_password = self._encode_login_data(password)
            
            # 使用加密后的数据
            from urllib.parse import urlencode
            login_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,      # 使用加密后的邮箱
                'password': encoded_password,  # 使用加密后的密码
                'fixed_mix_mode': '1'
            })
            
            login_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': self.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home?type=image',
                'user-agent': self.user_agent,
            }
            
            # 尝试获取 csrf token
            csrf_token = None
            for cookie in self.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            if csrf_token:
                login_headers['x-tt-passport-csrf-token'] = csrf_token
            
            response = self.session.post(
                login_url,
                params=login_params,
                headers=login_headers,
                data=login_data,
                timeout=15
            )
            
            response.raise_for_status()
            result = response.json()
            
            # 步骤3: 提交同意协议
            if result.get('message') == 'success':
                self._submit_consent()
            
            # 从响应中提取 session id
            if result.get('message') == 'success' and 'data' in result:
                sessionid = None
                for cookie in self.session.cookies:
                    if cookie.name in ['sessionid', 'sessionid_ss', 'sid_tt']:
                        sessionid = cookie.value
                        break
                
                if sessionid:
                    result['extracted_sessionid'] = sessionid
                    
                    # 步骤4: 获取完整用户信息
                    user_info = self.get_detailed_user_info()
                    if user_info.get('ret') == '0':
                        result['detailed_user_info'] = user_info.get('data', {})
                    
                    return result
                else:
                    result['warning'] = '登录成功但未能提取到 sessionid，请查看完整响应'
                    return result
            else:
                return result
                
        except requests.exceptions.RequestException as e:
            error_details = ""
            if hasattr(e, 'response') and e.response is not None:
                try: 
                    error_details = e.response.json()
                except json.JSONDecodeError: 
                    error_details = e.response.text
            return {"error": f"登录请求失败: {e}", "details": error_details}
    
    def _submit_consent(self):
        """提交用户同意协议"""
        try:
            consent_url = "https://dreamina.capcut.com/lv/v1/sc/compliance_popup/submit_consent"
            
            consent_data = {
                "entity_key": "conditions-policy-privacy-policy",
                "business_flow": "web_login",
                "consent_status": "approved",
                "region": "NG"
            }
            
            headers = {
                'accept': 'application/json, text/plain, */*',
                'content-type': 'application/json',
                'did': self.did,
                'appid': '513641',
                'pf': self.pf,
                'lan': 'en',
                'tdid': '0',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home?type=image',
                'user-agent': self.user_agent,
            }
            
            self.session.post(
                consent_url,
                headers=headers,
                json=consent_data,
                timeout=10
            )
        except:
            pass  # 同意协议失败不影响登录
    
    def get_detailed_user_info(self):
        """获取详细的用户信息（包含空间信息、订阅状态等）"""
        api_url = "https://dreamina.capcut.com/lv/v1/user/web/user_info"
        
        payload = {
            "sem_info": {
                "is_sem": False,
                "medium": "Direct",
                "register_source": "direct",
                "register_second_source": "enter_url"
            }
        }
        
        payload_str = json.dumps(payload)
        
        # 使用签名请求
        return self._send_request(api_url, None, payload_str, tdid_for_sign="", timeout=10)

    def send_register_code(self, email, password):
        """
        发送注册验证码到邮箱
        返回: {"email": "部分隐藏的邮箱", "email_ticket": "票据"}
        """
        verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
        
        try:
            # 步骤1: 地区验证
            region_url = "https://login-row.www.capcut.com/passport/web/region/"
            region_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp,
                'mix_mode': '1'
            }
            
            # 编码邮箱和临时ID
            encoded_email = self._encode_login_data(email)
            temp_id = self._encode_login_data(f"temp_{email}")
            hashed_id = hashlib.sha256(email.encode()).hexdigest()
            
            region_data = f"type=2&hashed_id={hashed_id}&temporary_id={temp_id}"
            
            region_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': self.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/',
                'user-agent': self.user_agent,
            }
            
            region_response = self.session.post(
                region_url,
                params=region_params,
                headers=region_headers,
                data=region_data,
                timeout=10
            )
            
            # 步骤2: 发送验证码
            send_code_url = "https://dreamina.capcut.com/passport/web/email/send_code/"
            send_code_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            # 对密码也进行加密
            encoded_password = self._encode_login_data(password)
            
            from urllib.parse import urlencode
            send_code_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,
                'password': encoded_password,
                'type': '34',  # 34 = 注册
                'fixed_mix_mode': '1'
            })
            
            # 获取 CSRF token
            csrf_token = None
            for cookie in self.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            send_code_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': self.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home',
                'user-agent': self.user_agent,
            }
            
            if csrf_token:
                send_code_headers['x-tt-passport-csrf-token'] = csrf_token
            
            response = self.session.post(
                send_code_url,
                params=send_code_params,
                headers=send_code_headers,
                data=send_code_data,
                timeout=15
            )
            
            response.raise_for_status()
            result = response.json()
            
            if result.get('message') == 'success':
                return {
                    'success': True,
                    'email': result.get('data', {}).get('email', email),
                    'email_ticket': result.get('data', {}).get('email_ticket', ''),
                    'message': f"验证码已发送到 {result.get('data', {}).get('email', email)}"
                }
            else:
                return {
                    'success': False,
                    'error': result.get('data', {}).get('description', '发送验证码失败')
                }
                
        except requests.exceptions.RequestException as e:
            error_details = ""
            if e.response:
                try:
                    error_details = e.response.json()
                except json.JSONDecodeError:
                    error_details = e.response.text
            return {
                'success': False,
                'error': f"请求失败: {e}",
                'details': error_details
            }
    
    def get_temp_email(self):
        """
        生成临时邮箱地址（使用 tempmail.plus 的 mailto.plus 域名）
        返回: 邮箱地址字符串，例如 "abc123@mailto.plus"
        """
        import random
        import string
        
        # 生成随机邮箱前缀（5位随机字符）
        prefix = ''.join(random.choices(string.ascii_lowercase + string.digits, k=5))
        email = f"{prefix}@mailto.plus"
        
        return email
    
    def check_verification_code(self, email, timeout=120):
        """
        轮询检查临时邮箱中的验证码
        email: 临时邮箱地址
        timeout: 超时时间（秒）
        返回: {"success": True, "code": "ABC123"} 或 {"success": False, "error": "..."}
        """
        import time
        import re
        from urllib.parse import quote
        
        start_time = time.time()
        encoded_email = quote(email)
        
        # 轮询检查邮件
        while time.time() - start_time < timeout:
            try:
                # 获取邮件列表
                list_url = f"https://tempmail.plus/api/mails?email={encoded_email}&first_id=0&epin="
                list_headers = {
                    'accept': 'application/json, text/javascript, */*; q=0.01',
                    'accept-language': 'zh-CN,zh;q=0.9',
                    'user-agent': self.user_agent,
                    'x-requested-with': 'XMLHttpRequest',
                    'referer': 'https://tempmail.plus/zh/'
                }
                
                list_response = self.session.get(list_url, headers=list_headers, timeout=10)
                list_response.raise_for_status()
                list_data = list_response.json()
                
                # 检查是否有邮件
                if list_data.get('result') and list_data.get('count', 0) > 0:
                    mail_list = list_data.get('mail_list', [])
                    
                    # 查找来自 Dreamina 的邮件
                    for mail in mail_list:
                        if 'dreamina@mail.capcut.com' in mail.get('from_mail', '').lower():
                            subject = mail.get('subject', '')
                            mail_id = mail.get('mail_id')
                            
                            # 尝试从标题提取验证码
                            code_match = re.search(r'verification code is ([A-Z0-9]{6})', subject)
                            if code_match:
                                return {
                                    'success': True,
                                    'code': code_match.group(1),
                                    'mail_id': mail_id
                                }
                            
                            # 如果标题没有，获取邮件详情
                            detail_url = f"https://tempmail.plus/api/mails/{mail_id}?email={encoded_email}&epin="
                            detail_response = self.session.get(detail_url, headers=list_headers, timeout=10)
                            detail_response.raise_for_status()
                            detail_data = detail_response.json()
                            
                            if detail_data.get('result'):
                                # 从文本内容提取验证码
                                text_content = detail_data.get('text', '')
                                html_content = detail_data.get('html', '')
                                
                                # 匹配各种格式的验证码
                                patterns = [
                                    r'verification code:\s*([A-Z0-9]{6})',
                                    r'code is\s*([A-Z0-9]{6})',
                                    r'验证码[：:]\s*([A-Z0-9]{6})',
                                    r'<span[^>]*>([A-Z0-9]{6})</span>'
                                ]
                                
                                for pattern in patterns:
                                    match = re.search(pattern, text_content + html_content, re.IGNORECASE)
                                    if match:
                                        return {
                                            'success': True,
                                            'code': match.group(1),
                                            'mail_id': mail_id
                                        }
                
                # 等待5秒后重试
                time.sleep(5)
                
            except Exception as e:
                # 忽略单次请求错误，继续轮询
                time.sleep(5)
                continue
        
        # 超时
        return {
            'success': False,
            'error': f'在 {timeout} 秒内未收到验证码邮件'
        }
    
    def register_with_email(self, email, password, code, email_ticket, birthday, region='JP'):
        """
        使用邮箱验证码完成注册
        email: 原始邮箱
        password: 原始密码
        code: 从邮箱收到的验证码
        email_ticket: 发送验证码时获取的票据
        birthday: 生日，格式 YYYY-MM-DD
        region: 地区代码，默认 JP
        返回: 用户信息和自动登录的session
        """
        verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
        
        try:
            # 验证码验证并注册（包含生日）
            register_url = "https://dreamina.capcut.com/passport/web/email/register_verify_login/"
            register_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            # 加密邮箱、密码和验证码
            encoded_email = self._encode_login_data(email)
            encoded_password = self._encode_login_data(password)
            encoded_code = self._encode_login_data(code)
            
            from urllib.parse import urlencode
            register_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,
                'code': encoded_code,
                'password': encoded_password,
                'type': '34',
                'email_ticket': email_ticket,
                'birthday': birthday,
                'force_user_region': region,
                'biz_param': '{}',
                'fixed_mix_mode': '1'
            })
            
            # 获取 CSRF token
            csrf_token = None
            for cookie in self.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            register_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': self.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home',
                'user-agent': self.user_agent,
            }
            
            if csrf_token:
                register_headers['x-tt-passport-csrf-token'] = csrf_token
            
            response = self.session.post(
                register_url,
                params=register_params,
                headers=register_headers,
                data=register_data,
                timeout=15
            )
            
            response.raise_for_status()
            result = response.json()
            
            if result.get('message') == 'success':
                # 注册成功，自动登录，提取 session
                session_id = None
                for cookie in self.session.cookies:
                    if cookie.name == 'sessionid':
                        session_id = cookie.value
                        break
                
                user_data = result.get('data', {})
                
                # 提交用户协议
                try:
                    self._submit_consent()
                except:
                    pass  # 协议提交失败不影响注册
                
                return {
                    'success': True,
                    'session_id': session_id,
                    'user_info': user_data,
                    'message': f"注册成功！用户: {user_data.get('screen_name', 'unknown')}"
                }
            else:
                return {
                    'success': False,
                    'error': result.get('data', {}).get('description', '注册失败')
                }
                
        except requests.exceptions.RequestException as e:
            error_details = ""
            if e.response:
                try:
                    error_details = e.response.json()
                except json.JSONDecodeError:
                    error_details = e.response.text
            return {
                'success': False,
                'error': f"请求失败: {e}",
                'details': error_details
            }

    # --- 已修改: 添加了 model_key 参数 ---
    def create_text_to_image_task(self, prompt, negative_prompt="", model_key="high_aes_general_v30l:general_v3.0_18b"):
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/aigc_draft/generate"
        # --- 已修改: 使用 [MODEL_KEY] 作为占位符，更新到 3.3.2 版本 ---
        payload_template = '{"extend":{"root_model":"[MODEL_KEY]"},"submit_id":"[SUBMIT_ID]","metrics_extra":"{\\"promptSource\\":\\"custom\\",\\"generateCount\\":1,\\"enterFrom\\":\\"click\\",\\"generateId\\":\\"[SUBMIT_ID]\\",\\"isRegenerate\\":false}","draft_content":"{\\"type\\":\\"draft\\",\\"id\\":\\"[DRAFT_ID]\\",\\"min_version\\":\\"3.0.2\\",\\"min_features\\":[],\\"is_from_tsn\\":true,\\"version\\":\\"3.3.2\\",\\"main_component_id\\":\\"[COMPONENT_ID]\\",\\"component_list\\":[{\\"type\\":\\"image_base_component\\",\\"id\\":\\"[COMPONENT_ID]\\",\\"min_version\\":\\"3.0.2\\",\\"aigc_mode\\":\\"workbench\\",\\"gen_type\\":1,\\"metadata\\":{\\"type\\":\\"\\",\\"id\\":\\"[METADATA_ID]\\",\\"created_platform\\":3,\\"created_platform_version\\":\\"\\",\\"created_time_in_ms\\":\\"[TIMESTAMP_MS]\\",\\"created_did\\":\\"\\"},\\"generate_type\\":\\"generate\\",\\"abilities\\":{\\"type\\":\\"\\",\\"id\\":\\"[ABILITIES_ID]\\",\\"generate\\":{\\"type\\":\\"\\",\\"id\\":\\"[GENERATE_ID]\\",\\"core_param\\":{\\"type\\":\\"\\",\\"id\\":\\"[CORE_PARAM_ID]\\",\\"model\\":\\"[MODEL_KEY]\\",\\"prompt\\":\\"[PROMPT]\\",\\"negative_prompt\\":\\"[NEGATIVE_PROMPT]\\",\\"seed\\":[SEED],\\"sample_strength\\":0.5,\\"image_ratio\\":1,\\"large_image_info\\":{\\"type\\":\\"\\",\\"id\\":\\"[LARGE_IMAGE_ID]\\",\\"height\\":1328,\\"width\\":1328,\\"resolution_type\\":\\"1k\\"},\\"intelligent_ratio\\":false}}}}]}","http_common_info":{"aid":513641}}'
        current_timestamp_ms = str(int(time.time() * 1000))
        submit_id = str(uuid.uuid4())
        # --- 已修改: 替换 [MODEL_KEY] 占位符 ---
        payload_str = payload_template.replace("[MODEL_KEY]", model_key).replace("[SUBMIT_ID]", submit_id).replace("[DRAFT_ID]", str(uuid.uuid4())).replace("[COMPONENT_ID]", str(uuid.uuid4())).replace("[METADATA_ID]", str(uuid.uuid4())).replace("[TIMESTAMP_MS]", current_timestamp_ms).replace("[ABILITIES_ID]", str(uuid.uuid4())).replace("[GENERATE_ID]", str(uuid.uuid4())).replace("[CORE_PARAM_ID]", str(uuid.uuid4())).replace("[PROMPT]", json.dumps(prompt, ensure_ascii=False)[1:-1]).replace("[NEGATIVE_PROMPT]", json.dumps(negative_prompt, ensure_ascii=False)[1:-1]).replace("[SEED]", str(random.randint(0, 999999999))).replace("[LARGE_IMAGE_ID]", str(uuid.uuid4()))
        return self._send_request(api_url, self.common_url_params, payload_str, tdid_for_sign="web")

    def create_text_to_video_task(self, prompt, ratio="1:1", duration_ms=5000):
        # ... 此函数及后续函数保持不变 ...
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/aigc_draft/generate"
        payload_template = '{"extend":{"root_model":"dreamina_ic_generate_video_model_vgfm_3.0","m_video_commerce_info":{"benefit_type":"basic_video_operation_vgfm_v_three","resource_id":"generate_video","resource_id_type":"str","resource_sub_type":"aigc"},"m_video_commerce_info_list":[{"benefit_type":"basic_video_operation_vgfm_v_three","resource_id":"generate_video","resource_id_type":"str","resource_sub_type":"aigc"}]},"submit_id":"[SUBMIT_ID]","metrics_extra":"{\\"promptSource\\":\\"custom\\",\\"isDefaultSeed\\":1,\\"originSubmitId\\":\\"[SUBMIT_ID]\\",\\"isRegenerate\\":false,\\"enterFrom\\":\\"click\\"}","draft_content":"{\\"type\\":\\"draft\\",\\"id\\":\\"[DRAFT_ID]\\",\\"min_version\\":\\"3.0.5\\",\\"min_features\\":[],\\"is_from_tsn\\":true,\\"version\\":\\"3.2.8\\",\\"main_component_id\\":\\"[COMPONENT_ID]\\",\\"component_list\\":[{\\"type\\":\\"video_base_component\\",\\"id\\":\\"[COMPONENT_ID]\\",\\"min_version\\":\\"1.0.0\\",\\"aigc_mode\\":\\"workbench\\",\\"metadata\\":{\\"type\\":\\"\\",\\"id\\":\\"[METADATA_ID]\\",\\"created_platform\\":3,\\"created_platform_version\\":\\"\\",\\"created_time_in_ms\\":\\"[TIMESTAMP_MS]\\",\\"created_did\\":\\"\\"},\\"generate_type\\":\\"gen_video\\",\\"abilities\\":{\\"type\\":\\"\\",\\"id\\":\\"[ABILITIES_ID]\\",\\"gen_video\\":{\\"type\\":\\"\\",\\"id\\":\\"[GENERATE_ID]\\",\\"text_to_video_params\\":{\\"type\\":\\"\\",\\"id\\":\\"[T2V_PARAM_ID]\\",\\"video_gen_inputs\\":[{\\"type\\":\\"\\",\\"id\\":\\"[INPUT_ID]\\",\\"min_version\\":\\"3.0.5\\",\\"prompt\\":\\"[PROMPT]\\",\\"video_mode\\":2,\\"fps\\":24,\\"duration_ms\\":[DURATION_MS],\\"resolution\\":\\"720p\\"}],\\"video_aspect_ratio\\":\\"[RATIO]\\",\\"seed\\":[SEED],\\"model_req_key\\":\\"dreamina_ic_generate_video_model_vgfm_3.0\\",\\"priority\\":0},\\"video_task_extra\\":\\"{\\\\\\"promptSource\\\\\\":\\\\\\"custom\\\\\\",\\\\\\"isDefaultSeed\\\\\\":1,\\\\\\"originSubmitId\\\\\\":\\\\\\"[SUBMIT_ID]\\\\\\",\\\\\\"isRegenerate\\\\\\":false,\\\\\\"enterFrom\\\\\\":\\\\\\"click\\\\\\"}\\"}},\\"process_type\\":1}]}","http_common_info":{"aid":513641}}'
        submit_id = str(uuid.uuid4())
        current_timestamp_ms = str(int(time.time() * 1000))
        payload_str = payload_template.replace("[SUBMIT_ID]", submit_id).replace("[DRAFT_ID]", str(uuid.uuid4())).replace("[COMPONENT_ID]", str(uuid.uuid4())).replace("[METADATA_ID]", str(uuid.uuid4())).replace("[TIMESTAMP_MS]", current_timestamp_ms).replace("[ABILITIES_ID]", str(uuid.uuid4())).replace("[GENERATE_ID]", str(uuid.uuid4())).replace("[T2V_PARAM_ID]", str(uuid.uuid4())).replace("[INPUT_ID]", str(uuid.uuid4())).replace("[PROMPT]", json.dumps(prompt, ensure_ascii=False)[1:-1]).replace("[RATIO]", json.dumps(ratio, ensure_ascii=False)[1:-1]).replace("[DURATION_MS]", str(duration_ms)).replace("[SEED]", str(random.randint(0, 4294967295)))
        return self._send_request(api_url, self.common_url_params, payload_str, tdid_for_sign="web")

    def query_video_task_status(self, submit_id):
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/get_history_by_ids"
        payload_str = json.dumps({"submit_ids": [submit_id]})
        return self._send_request(api_url, self.common_url_params, payload_str, tdid_for_sign="web")

    def get_upload_token(self, scene=2):
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/get_upload_token"
        payload = {"scene": scene}
        url_params = self.common_url_params.copy()
        url_params['da_version'] = '3.3.0' 
        return self._send_request(api_url, url_params, json.dumps(payload), tdid_for_sign="web")

    def apply_image_upload(self, file_size_bytes, upload_data):
        upload_domain = upload_data.get('upload_domain')
        space_name = upload_data.get('space_name')
        if not upload_domain or not space_name:
            raise ValueError("apply_image_upload: 'upload_domain' or 'space_name' is missing from token data")

        api_url = f"https://{upload_domain}/"
        params = {
            'Action': 'ApplyImageUpload',
            'Version': '2018-08-01',
            'ServiceId': space_name,
            'FileSize': file_size_bytes
        }
        
        try:
            headers = self._build_aws4_headers(upload_data, api_url, params, 'GET')
            response = self.session.get(api_url, params=params, headers=headers, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            details = getattr(e, 'response', None)
            details_text = details.text if details is not None else "No response"
            return {"error": f"申请上传失败: {e}", "details": details_text}

    def upload_image_file(self, image_bytes, upload_info):
        upload_url = f"https://{upload_info['UploadHost']}/{upload_info['StoreUri']}"
        headers = {
            'Authorization': upload_info['Auth'],
            'Content-Type': 'application/octet-stream',
            'Content-Disposition': 'attachment; filename="undefined"'
        }
        crc = zlib.crc32(image_bytes)
        headers['Content-CRC32'] = f'{crc:x}'
        
        try:
            response = self.session.post(upload_url, data=image_bytes, headers=headers, timeout=60)
            response.raise_for_status()
            return {"code": 2000, "message": "Upload successful"}
        except requests.exceptions.RequestException as e:
            return {"error": f"上传文件失败: {e}"}

    def commit_image_upload(self, session_key, upload_data):
        upload_domain = upload_data.get('upload_domain')
        space_name = upload_data.get('space_name')
        if not upload_domain or not space_name:
            raise ValueError("commit_image_upload: 'upload_domain' or 'space_name' is missing")

        api_url = f"https://{upload_domain}/"
        params = {'Action': 'CommitImageUpload', 'Version': '2018-08-01', 'ServiceId': space_name}
        payload_str = json.dumps({"SessionKey": session_key})
        
        try:
            headers = self._build_aws4_headers(upload_data, api_url, params, 'POST', payload_str)
            response = self.session.post(api_url, params=params, data=payload_str, headers=headers, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            details = getattr(e, 'response', None)
            details_text = details.text if details is not None else "No response"
            return {"error": f"确认上传失败: {e}", "details": details_text}

    def _build_aws4_headers(self, upload_credentials: dict, api_url: str, params: dict, method: str, payload_str: str = '') -> dict:
        access_key_id = upload_credentials.get("access_key_id")
        secret_access_key = upload_credentials.get("secret_access_key")
        session_token = upload_credentials.get("session_token")
        region = upload_credentials.get("region", "sg")
        service = 'imagex'
        
        if not all([access_key_id, secret_access_key, session_token]):
            raise ValueError("缺少必要的AWS凭据")
        
        parsed_url = urlparse(api_url)
        host = parsed_url.netloc
        path = parsed_url.path or '/'
        
        now = datetime.utcnow()
        amz_date = now.strftime('%Y%m%dT%H%M%SZ')
        date_stamp = now.strftime('%Y%m%d')
        
        query_items = []
        for key in sorted(params.keys()):
            value = str(params[key])
            query_items.append(f"{urlencode({key: value})}")
        canonical_querystring = '&'.join(query_items).replace('+', '%20')

        if payload_str:
            canonical_headers = f'content-type:application/json\nhost:{host}\nx-amz-date:{amz_date}\nx-amz-security-token:{session_token}\n'
            signed_headers = 'content-type;host;x-amz-date;x-amz-security-token'
        else:
            canonical_headers = f'host:{host}\nx-amz-date:{amz_date}\nx-amz-security-token:{session_token}\n'
            signed_headers = 'host;x-amz-date;x-amz-security-token'

        payload_hash = hashlib.sha256(payload_str.encode('utf-8')).hexdigest()
        
        canonical_request = f"{method}\n{path}\n{canonical_querystring}\n{canonical_headers}\n{signed_headers}\n{payload_hash}"
        
        algorithm = 'AWS4-HMAC-SHA256'
        credential_scope = f"{date_stamp}/{region}/{service}/aws4_request"
        string_to_sign = f"{algorithm}\n{amz_date}\n{credential_scope}\n{hashlib.sha256(canonical_request.encode('utf-8')).hexdigest()}"
        
        def sign(key, msg):
            return hmac.new(key, msg.encode('utf-8'), hashlib.sha256).digest()
        
        k_date = sign(('AWS4' + secret_access_key).encode('utf-8'), date_stamp)
        k_region = sign(k_date, region)
        k_service = sign(k_region, service)
        k_signing = sign(k_service, 'aws4_request')
        
        signature = hmac.new(k_signing, string_to_sign.encode('utf-8'), hashlib.sha256).hexdigest()
        
        authorization_header = f"{algorithm} Credential={access_key_id}/{credential_scope}, SignedHeaders={signed_headers}, Signature={signature}"
        
        final_headers = {
            'Authorization': authorization_header,
            'X-Amz-Date': amz_date,
            'X-Amz-Security-Token': session_token,
            'User-Agent': self.user_agent,
        }
        if payload_str:
            final_headers['Content-Type'] = 'application/json'

        return final_headers

    def create_image_to_video_task(self, prompt, start_frame_uri, ratio="16:9", resolution="720p", duration_ms=5000, model="Video 3.0", use_first_last=False, end_frame_uri=None, ending_control=1.0):
        api_url = "https://mweb-api-sg.capcut.com/mweb/v1/aigc_draft/generate"
        
        final_url_params = self.common_url_params.copy()
        final_url_params['da_version'] = '3.3.0'

        submit_id = str(uuid.uuid4())
        draft_id, component_id, metadata_id, abilities_id, gen_video_id, t2v_param_id, input_id = (str(uuid.uuid4()) for _ in range(7))
        current_timestamp_ms = str(int(time.time() * 1000))

        metrics_extra_data = {
            "promptSource": "custom", "isDefaultSeed": 1, "originSubmitId": submit_id,
            "isRegenerate": False, "enterFrom": "click"
        }
        if use_first_last and end_frame_uri:
            metrics_extra_data["functionMode"] = "first_last_frames"

        video_gen_inputs_item = {
            "type": "", "id": input_id, "min_version": "3.0.5", "prompt": prompt,
            "first_frame_image": {
                "type": "image", "id": str(uuid.uuid4()), "source_from": "upload", "platform_type": 1, "name": "",
                "image_uri": start_frame_uri, "width": 1024, "height": 1024, "format": "", "uri": start_frame_uri
            },
            "video_mode": 2, "fps": 24, "duration_ms": duration_ms, "resolution": resolution
        }

        if use_first_last and end_frame_uri:
            video_gen_inputs_item['end_frame_image'] = {
                "type": "image", "id": str(uuid.uuid4()), "source_from": "upload", "platform_type": 1, "name": "",
                "image_uri": end_frame_uri, "width": 1024, "height": 1024, "format": "", "uri": end_frame_uri
            }
            video_gen_inputs_item['ending_control'] = str(float(ending_control))

        payload = {
            "extend": { "root_model": "dreamina_ic_generate_video_model_vgfm_3.0", "m_video_commerce_info": { "benefit_type": "basic_video_operation_vgfm_v_three", "resource_id": "generate_video", "resource_id_type": "str", "resource_sub_type": "aigc" }, "m_video_commerce_info_list": [{ "benefit_type": "basic_video_operation_vgfm_v_three", "resource_id": "generate_video", "resource_id_type": "str", "resource_sub_type": "aigc" }] },
            "submit_id": submit_id,
            "metrics_extra": json.dumps(metrics_extra_data),
            "draft_content": json.dumps({
                "type": "draft", "id": draft_id, "min_version": "3.0.5", "min_features": [], "is_from_tsn": True, "version": "3.3.0", "main_component_id": component_id,
                "component_list": [{
                    "type": "video_base_component", "id": component_id, "min_version": "1.0.0", "aigc_mode": "workbench",
                    "metadata": { "type": "", "id": metadata_id, "created_platform": 3, "created_platform_version": "", "created_time_in_ms": current_timestamp_ms, "created_did": "" },
                    "generate_type": "gen_video",
                    "abilities": { "type": "", "id": abilities_id,
                        "gen_video": { "type": "", "id": gen_video_id,
                            "text_to_video_params": {
                                "type": "", "id": t2v_param_id,
                                "video_gen_inputs": [video_gen_inputs_item],
                                "video_aspect_ratio": ratio, "seed": random.randint(0, 4294967295), "model_req_key": "dreamina_ic_generate_video_model_vgfm_3.0", "priority": 0
                            },
                            "video_task_extra": json.dumps(metrics_extra_data)
                        }
                    },
                    "process_type": 1
                }]
            }),
            "http_common_info": {"aid": 513641}
        }
        
        return self._send_request(api_url, final_url_params, json.dumps(payload), tdid_for_sign="web")

    def query_task_status(self, submit_id):
        return self.query_video_task_status(submit_id)

# ==============================================================================
# 账号数据库管理类
# ==============================================================================
class AccountDatabase:
    """账号数据库管理器"""
    
    def __init__(self, db_file="accounts_database.json"):
        self.db_file = db_file
        self.accounts = []
        self.load_database()
    
    def load_database(self):
        """从文件加载数据库"""
        try:
            if os.path.exists(self.db_file):
                with open(self.db_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.accounts = data.get('accounts', [])
        except Exception as e:
            print(f"加载数据库失败: {e}")
            self.accounts = []
    
    def save_database(self):
        """保存数据库到文件"""
        try:
            data = {
                'accounts': self.accounts,
                'statistics': self.get_statistics(),
                'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            with open(self.db_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            print(f"保存数据库失败: {e}")
            return False
    
    def add_account(self, email, password, session_id, did, user_info=None):
        """添加新账号"""
        account_id = len(self.accounts) + 1
        account = {
            'id': account_id,
            'email': email,
            'password': password,
            'session_id': session_id,
            'did': did,
            'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'status': 'active',
            'user_info': user_info or {}
        }
        self.accounts.append(account)
        self.save_database()
        return account
    
    def delete_account(self, account_id):
        """删除账号"""
        self.accounts = [acc for acc in self.accounts if acc['id'] != account_id]
        self.save_database()
    
    def delete_selected_accounts(self, account_ids):
        """批量删除账号"""
        self.accounts = [acc for acc in self.accounts if acc['id'] not in account_ids]
        self.save_database()
    
    def clear_all(self):
        """清空所有账号"""
        self.accounts = []
        self.save_database()
    
    def get_all_accounts(self):
        """获取所有账号"""
        return self.accounts
    
    def get_statistics(self):
        """获取统计信息"""
        total = len(self.accounts)
        active = len([acc for acc in self.accounts if acc['status'] == 'active'])
        failed = len([acc for acc in self.accounts if acc['status'] == 'failed'])
        return {
            'total': total,
            'active': active,
            'failed': failed
        }
    
    def export_to_csv(self, filepath):
        """导出为CSV文件"""
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['序号', '邮箱', '密码', 'SessionID', 'DID', '创建时间', '状态'])
                for acc in self.accounts:
                    status = '✅活跃' if acc['status'] == 'active' else '❌失效'
                    writer.writerow([
                        acc['id'],
                        acc['email'],
                        acc['password'],
                        acc['session_id'],
                        acc['did'],
                        acc['create_time'],
                        status
                    ])
            return True
        except Exception as e:
            print(f"导出CSV失败: {e}")
            return False
    
    def export_to_excel(self, filepath):
        """导出为Excel文件（需要openpyxl库）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "账号列表"
            
            # 设置表头
            headers = ['序号', '邮箱', '密码', 'SessionID', 'DID', '创建时间', '状态']
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加数据
            for acc in self.accounts:
                status = '✅活跃' if acc['status'] == 'active' else '❌失效'
                ws.append([
                    acc['id'],
                    acc['email'],
                    acc['password'],
                    acc['session_id'],
                    acc['did'],
                    acc['create_time'],
                    status
                ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 10
            
            # 添加统计信息到第二个sheet
            ws2 = wb.create_sheet("统计信息")
            stats = self.get_statistics()
            ws2.append(['统计项', '数量'])
            ws2.append(['总账号数', stats['total']])
            ws2.append(['活跃账号', stats['active']])
            ws2.append(['失效账号', stats['failed']])
            
            wb.save(filepath)
            return True
        except ImportError:
            # 如果没有openpyxl，降级为CSV
            return self.export_to_csv(filepath.replace('.xlsx', '.csv'))
        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False

# ==============================================================================
# 批量注册管理器
# ==============================================================================
class BatchRegisterManager:
    """批量并发注册管理器"""
    
    def __init__(self, count, concurrent, password, database, delay=8, progress_callback=None):
        self.count = count
        self.concurrent = concurrent
        self.password = password
        self.database = database
        self.delay = delay  # 每个账号之间的延迟时间（秒）
        self.progress_callback = progress_callback
        self.success_count = 0
        self.failed_count = 0
        self.stopped = False
        self.paused = False
        self.lock = threading.Lock()
    
    def stop(self):
        """停止注册"""
        self.stopped = True
    
    def pause(self):
        """暂停注册"""
        self.paused = True
    
    def resume(self):
        """恢复注册"""
        self.paused = False
    
    def register_one(self, index):
        """注册单个账号"""
        if self.stopped:
            return None
        
        # 等待暂停
        while self.paused and not self.stopped:
            time.sleep(0.5)
        
        if self.stopped:
            return None
        
        # ========== 添加随机延迟，避免触发限制 ==========
        # 每个账号注册前等待指定时间（±2秒的随机浮动）
        delay = self.delay + random.uniform(-2, 2)
        delay = max(3, delay)  # 至少3秒
        time.sleep(delay)
        
        # 重试机制（最多3次）
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            if self.stopped:
                return None
            
            try:
                # 创建 API 实例
                api = CapCutAPI("", "")
                
                # 生成临时邮箱
                email = api.get_temp_email()
                
                # 发送验证码（如果失败会抛出异常）
                send_result = api.send_register_code(email, self.password)
                if not send_result.get('success'):
                    error_msg = send_result.get('error', '')
                    
                    # 检查是否是频率限制错误
                    if 'Maximum number' in error_msg or 'Try again later' in error_msg:
                        retry_count += 1
                        if retry_count < max_retries:
                            # 等待更长时间后重试（指数退避）
                            wait_time = 30 * (2 ** (retry_count - 1))  # 30秒, 60秒, 120秒
                            if self.progress_callback:
                                self.progress_callback('log', index, f"触发频率限制，等待 {wait_time} 秒后重试 ({retry_count}/{max_retries})...")
                            time.sleep(wait_time)
                            continue
                    
                    raise Exception(f"发送验证码失败: {error_msg}")
                
                # 发送成功，跳出重试循环
                break
                
            except Exception as e:
                retry_count += 1
                if retry_count >= max_retries:
                    # 达到最大重试次数，抛出异常
                    raise Exception(f"重试 {max_retries} 次后仍失败: {str(e)}")
                
                # 等待后重试
                wait_time = 15 * retry_count
                if self.progress_callback:
                    self.progress_callback('log', index, f"发送失败，{wait_time}秒后重试 ({retry_count}/{max_retries})...")
                time.sleep(wait_time)
        
        # 继续完成注册流程
        try:
            # 自动获取验证码
            code_result = api.check_verification_code(email, timeout=120)
            if not code_result.get('success'):
                raise Exception(f"获取验证码失败: {code_result.get('error')}")
            
            code = code_result.get('code')
            
            # 验证验证码
            verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
            verify_url = "https://dreamina.capcut.com/passport/web/email/register/code_verify/"
            verify_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            encoded_email = api._encode_login_data(email)
            encoded_code = api._encode_login_data(code)
            
            from urllib.parse import urlencode
            verify_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,
                'code': encoded_code,
                'type': '34',
                'fixed_mix_mode': '1'
            })
            
            csrf_token = None
            for cookie in api.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            verify_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': api.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home',
                'user-agent': api.user_agent,
            }
            
            if csrf_token:
                verify_headers['x-tt-passport-csrf-token'] = csrf_token
            
            verify_response = api.session.post(
                verify_url,
                params=verify_params,
                headers=verify_headers,
                data=verify_data,
                timeout=15
            )
            verify_result = verify_response.json()
            
            if verify_result.get('message') != 'success':
                raise Exception(f"验证失败: {verify_result.get('data', {}).get('description')}")
            
            new_email_ticket = verify_result.get('data', {}).get('email_ticket', '')
            
            # 完成注册
            birthday = "2000-01-01"
            register_result = api.register_with_email(email, self.password, code, new_email_ticket, birthday)
            
            if not register_result.get('success'):
                raise Exception(f"注册失败: {register_result.get('error')}")
            
            # 保存到数据库
            session_id = register_result.get('session_id')
            user_info = register_result.get('user_info', {})
            
            account = self.database.add_account(
                email=email,
                password=self.password,
                session_id=session_id,
                did=api.did,
                user_info=user_info
            )
            
            with self.lock:
                self.success_count += 1
            
            if self.progress_callback:
                self.progress_callback('success', index, account)
            
            return account
            
        except Exception as e:
            with self.lock:
                self.failed_count += 1
            
            if self.progress_callback:
                self.progress_callback('failed', index, str(e))
            
            return None
    
    def batch_register(self):
        """批量注册主函数"""
        with ThreadPoolExecutor(max_workers=self.concurrent) as executor:
            futures = []
            for i in range(self.count):
                if self.stopped:
                    break
                future = executor.submit(self.register_one, i + 1)
                futures.append(future)
            
            for future in as_completed(futures):
                if self.stopped:
                    break
                try:
                    result = future.result()
                except Exception as e:
                    print(f"注册任务异常: {e}")
        
        if self.progress_callback:
            self.progress_callback('complete', self.count, None)

# ==============================================================================
# Tkinter UI 界面
# ==============================================================================
class Application(tk.Tk):
    CONFIG_FILE = "capcut_config.json"
    
    # --- 模型名称到API密钥的映射字典 ---
    # 通过 "获取配置" 按钮可以查看所有最新模型及其密钥
    MODEL_MAPPING = {
        "Image 4.0 🆕": "high_aes_general_v40",
        "Nano banana 🆕": "external_model_gemini_flash_image_v25",
        "Image 3.1": "high_aes_general_v30l_art:general_v3.0_18b",
        "Image 3.0 (默认)": "high_aes_general_v30l:general_v3.0_18b",
        "Image 2.1": "high_aes_general_v21_L:general_v2.1_L",
        "Image 2.0 Pro": "high_aes_general_v20_L:general_v2.0_L",
        "Image 1.4": "high_aes_v14_dreamina:general_v1.4",
    }
    
    def __init__(self):
        super().__init__()
        self.title("CapCut AI创作工具 (v9.0 - 账号管理专业版)")
        self.geometry("1200x900")  # 增加宽度以适应账号管理界面
        
        self.api_queue = queue.Queue()
        self.download_images_var = tk.BooleanVar(value=True)
        self.download_path = os.path.join("downloads", "images")  # 默认下载路径
        self.initial_wait_time = 10  # 默认初始等待时间（秒）
        
        # 初始化账号数据库
        self.account_db = AccountDatabase()
        self.batch_manager = None  # 批量注册管理器
        
        self.create_widgets()
        self._toggle_login_method()  # 初始化登录方式显示
        self.load_config()
        self.process_queue()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        config_frame = ttk.LabelFrame(main_frame, text="配置信息", padding="10")
        config_frame.pack(fill=tk.X, pady=5)
        
        # 登录方式选择
        login_method_frame = ttk.Frame(config_frame)
        login_method_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.login_method_var = tk.StringVar(value="cookie")
        ttk.Radiobutton(login_method_frame, text="使用 Cookie (从浏览器F12获取)", 
                       variable=self.login_method_var, value="cookie", 
                       command=self._toggle_login_method).pack(side=tk.LEFT)
        ttk.Radiobutton(login_method_frame, text="账号密码登录", 
                       variable=self.login_method_var, value="password",
                       command=self._toggle_login_method).pack(side=tk.LEFT, padx=10)
        
        # Cookie 输入区域
        self.cookie_frame = ttk.Frame(config_frame)
        self.cookie_frame.pack(fill=tk.X)
        ttk.Label(self.cookie_frame, text="Cookie 或 sessionid:").pack(anchor='w')
        self.cookie_text = tk.Text(self.cookie_frame, height=5, wrap='word')
        self.cookie_text.pack(fill=tk.X, expand=True)
        
        # 账号密码输入区域
        self.account_frame = ttk.Frame(config_frame)
        email_frame = ttk.Frame(self.account_frame)
        email_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(email_frame, text="邮箱账号:", width=10).pack(side=tk.LEFT)
        self.email_entry = ttk.Entry(email_frame)
        self.email_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        password_frame = ttk.Frame(self.account_frame)
        password_frame.pack(fill=tk.X)
        ttk.Label(password_frame, text="密码:", width=10).pack(side=tk.LEFT)
        self.password_entry = ttk.Entry(password_frame, show="*")
        self.password_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        login_btn_frame = ttk.Frame(self.account_frame)
        login_btn_frame.pack(fill=tk.X, pady=(10, 0))
        self.login_button = ttk.Button(login_btn_frame, text="🔐 登录获取Session", 
                                       command=self.start_login, style="Accent.TButton")
        self.login_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.register_button = ttk.Button(login_btn_frame, text="📝 注册新账号", 
                                         command=self.start_register)
        self.register_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.auto_register_button = ttk.Button(login_btn_frame, text="🤖 全自动注册", 
                                              command=self.start_auto_register, style="Accent.TButton")
        self.auto_register_button.pack(side=tk.LEFT)
        
        # DID 输入
        ttk.Label(config_frame, text="DID (可留空):").pack(anchor='w', pady=(5,0))
        self.did_entry = ttk.Entry(config_frame)
        self.did_entry.pack(fill=tk.X, expand=True)
        
        button_frame = ttk.Frame(config_frame); button_frame.pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="保存配置", command=self.save_config).pack(side=tk.LEFT)
        ttk.Button(button_frame, text="加载配置", command=self.load_config).pack(side=tk.LEFT, padx=5)
        
        self.user_info_button = ttk.Button(button_frame, text="👤 获取用户信息", command=self.start_user_info_query)
        self.user_info_button.pack(side=tk.RIGHT, padx=5)
        
        self.config_button = ttk.Button(button_frame, text="🔧 获取配置", command=self.start_config_query)
        self.config_button.pack(side=tk.RIGHT, padx=5)
        
        self.receive_credit_button = ttk.Button(button_frame, text="🎁 领取积分", command=self.start_credit_receive)
        self.receive_credit_button.pack(side=tk.RIGHT, padx=5)
        
        self.credit_button = ttk.Button(button_frame, text="💰 查询积分", command=self.start_credit_query)
        self.credit_button.pack(side=tk.RIGHT)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        t2i_frame = ttk.Frame(notebook, padding="10")
        t2v_frame = ttk.Frame(notebook, padding="10")
        i2v_frame = ttk.Frame(notebook, padding="10")
        account_frame = ttk.Frame(notebook, padding="10")  # 新增账号管理Tab
        notebook.add(t2i_frame, text='文生图 (Text-to-Image)')
        notebook.add(t2v_frame, text='文生视频 (Text-to-Video)')
        notebook.add(i2v_frame, text='图生视频 (Image-to-Video)')
        notebook.add(account_frame, text='💼 账号管理')  # 新增账号管理Tab

        self.create_t2i_widgets(t2i_frame)
        self.create_t2v_widgets(t2v_frame)
        self.create_i2v_widgets(i2v_frame)
        self.create_account_widgets(account_frame)  # 创建账号管理界面

        log_frame = ttk.LabelFrame(main_frame, text="日志和结果", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.log_area = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)

    def create_t2i_widgets(self, parent_frame):
        # --- 新增: 模型选择下拉菜单 ---
        model_frame = ttk.Frame(parent_frame)
        model_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(model_frame, text="选择模型:").pack(side=tk.LEFT, anchor='w')
        self.t2i_model_selector = ttk.Combobox(
            model_frame, 
            values=list(self.MODEL_MAPPING.keys()), 
            state="readonly",
            width=25
        )
        self.t2i_model_selector.set("Image 3.0 (默认)")
        self.t2i_model_selector.pack(side=tk.LEFT, padx=5)
        # --- 模型选择结束 ---

        ttk.Label(parent_frame, text="提示词 (Prompt):").pack(anchor='w')
        self.t2i_prompt_text = tk.Text(parent_frame, height=5, wrap='word')
        self.t2i_prompt_text.pack(fill=tk.X, expand=True, pady=(0, 5))
        ttk.Label(parent_frame, text="反向提示词 (Negative Prompt, 可选):").pack(anchor='w')
        self.t2i_neg_prompt_text = tk.Text(parent_frame, height=3, wrap='word')
        self.t2i_neg_prompt_text.pack(fill=tk.X, expand=True)
        
        # 下载设置区域
        download_frame = ttk.Frame(parent_frame)
        download_frame.pack(pady=5, fill=tk.X)
        
        self.download_check = ttk.Checkbutton(
            download_frame,
            text="✅ 自动下载图片",
            variable=self.download_images_var
        )
        self.download_check.pack(side=tk.LEFT)
        
        ttk.Button(download_frame, text="📁 选择下载文件夹", command=self.choose_download_folder).pack(side=tk.LEFT, padx=10)
        
        self.download_path_label = ttk.Label(download_frame, text=f"保存到: {self.download_path}", foreground="gray")
        self.download_path_label.pack(side=tk.LEFT)
        
        # 等待时间设置区域
        wait_frame = ttk.Frame(parent_frame)
        wait_frame.pack(pady=5, fill=tk.X)
        
        ttk.Label(wait_frame, text="⏱️ 初始等待时间:").pack(side=tk.LEFT)
        self.wait_time_spinbox = ttk.Spinbox(wait_frame, from_=5, to=300, width=8)
        self.wait_time_spinbox.set(self.initial_wait_time)
        self.wait_time_spinbox.pack(side=tk.LEFT, padx=5)
        ttk.Label(wait_frame, text="秒 (某些模型需要更长时间)", foreground="gray").pack(side=tk.LEFT)

        self.t2i_button = ttk.Button(parent_frame, text="✨ 生成图片 ✨", command=self.start_t2i_generation, style="Accent.TButton")
        self.t2i_button.pack(pady=10)
        self.style = ttk.Style(self); self.style.configure("Accent.TButton", font=("Helvetica", 12, "bold"))

    def create_t2v_widgets(self, parent_frame):
        # ... 此函数及后续函数保持不变 ...
        ttk.Label(parent_frame, text="提示词 (Prompt):").pack(anchor='w')
        self.t2v_prompt_text = tk.Text(parent_frame, height=5, wrap='word')
        self.t2v_prompt_text.pack(fill=tk.X, expand=True, pady=(0, 5))
        options_frame = ttk.Frame(parent_frame)
        options_frame.pack(fill=tk.X, pady=5)
        ttk.Label(options_frame, text="视频比例:").pack(side=tk.LEFT)
        self.video_ratio = ttk.Combobox(options_frame, values=["1:1", "16:9", "9:16", "4:3", "3:4"], state="readonly")
        self.video_ratio.set("1:1")
        self.video_ratio.pack(side=tk.LEFT, padx=5)
        ttk.Label(options_frame, text="视频时长(秒):").pack(side=tk.LEFT, padx=(10, 0))
        self.duration_spinbox = ttk.Spinbox(options_frame, from_=1, to=15, increment=1, width=5)
        self.duration_spinbox.set("5")
        self.duration_spinbox.pack(side=tk.LEFT, padx=5)
        self.t2v_button = ttk.Button(parent_frame, text="🎬 生成视频 🎬", command=self.start_t2v_generation, style="Accent.TButton")
        self.t2v_button.pack(pady=10)

    def create_i2v_widgets(self, parent_frame):
        ttk.Label(parent_frame, text="提示词 (Prompt):").pack(anchor='w')
        self.i2v_prompt_text = tk.Text(parent_frame, height=4, wrap='word')
        self.i2v_prompt_text.pack(fill=tk.X, expand=True, pady=(0, 5))

        start_frame_ui = ttk.Frame(parent_frame)
        start_frame_ui.pack(fill=tk.X, pady=2)
        ttk.Label(start_frame_ui, text="起始帧图片:", width=12).pack(side=tk.LEFT)
        self.i2v_start_frame_path = ttk.Entry(start_frame_ui)
        self.i2v_start_frame_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(start_frame_ui, text="浏览...", command=lambda: self._browse_file(self.i2v_start_frame_path)).pack(side=tk.LEFT)

        self.use_first_last_var = tk.BooleanVar(value=False)
        self.use_first_last_check = ttk.Checkbutton(parent_frame, text="启用首尾帧模式 (End Frame)", variable=self.use_first_last_var, command=self._toggle_end_frame)
        self.use_first_last_check.pack(anchor='w', pady=5)
        
        end_frame_ui = ttk.Frame(parent_frame)
        end_frame_ui.pack(fill=tk.X, pady=2)
        self.i2v_end_frame_label = ttk.Label(end_frame_ui, text="结束帧图片:", width=12)
        self.i2v_end_frame_label.pack(side=tk.LEFT)
        self.i2v_end_frame_path = ttk.Entry(end_frame_ui)
        self.i2v_end_frame_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.i2v_end_frame_browse_btn = ttk.Button(end_frame_ui, text="浏览...", command=lambda: self._browse_file(self.i2v_end_frame_path))
        self.i2v_end_frame_browse_btn.pack(side=tk.LEFT)

        options_frame = ttk.Frame(parent_frame)
        options_frame.pack(fill=tk.X, pady=10)
        ttk.Label(options_frame, text="视频比例:").pack(side=tk.LEFT)
        self.i2v_video_ratio = ttk.Combobox(options_frame, values=["1:1", "16:9", "9:16", "4:3", "3:4"], state="readonly")
        self.i2v_video_ratio.set("16:9"); self.i2v_video_ratio.pack(side=tk.LEFT, padx=5)
        ttk.Label(options_frame, text="视频时长(秒):").pack(side=tk.LEFT, padx=(10, 0))
        self.i2v_duration_spinbox = ttk.Spinbox(options_frame, from_=1, to=15, increment=1, width=5)
        self.i2v_duration_spinbox.set("5"); self.i2v_duration_spinbox.pack(side=tk.LEFT, padx=5)

        # 新增：模式选择
        mode_frame = ttk.LabelFrame(parent_frame, text="执行模式", padding="10")
        mode_frame.pack(fill=tk.X, pady=10)
        
        self.i2v_mode_var = tk.StringVar(value="full")
        ttk.Radiobutton(mode_frame, text="完整流程（上传+提交+轮询+下载）", variable=self.i2v_mode_var, value="full").pack(anchor='w')
        ttk.Radiobutton(mode_frame, text="仅提交任务（返回submit_id，不等待）", variable=self.i2v_mode_var, value="submit_only").pack(anchor='w')
        ttk.Radiobutton(mode_frame, text="查询已有任务（输入submit_id）", variable=self.i2v_mode_var, value="query_only").pack(anchor='w')
        
        # 查询模式的 submit_id 输入框
        query_frame = ttk.Frame(mode_frame)
        query_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Label(query_frame, text="Submit ID:").pack(side=tk.LEFT)
        self.i2v_submit_id_entry = ttk.Entry(query_frame)
        self.i2v_submit_id_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        self.i2v_button = ttk.Button(parent_frame, text="🚀 从图片生成视频 🚀", command=self.start_i2v_generation, style="Accent.TButton")
        self.i2v_button.pack(pady=10)
        
        self._toggle_end_frame()
    
    def create_account_widgets(self, parent_frame):
        """创建账号管理界面"""
        # ========== 批量注册区域 ==========
        batch_frame = ttk.LabelFrame(parent_frame, text="🤖 批量自动注册", padding="15")
        batch_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 第一行：注册数量和并发数
        row1 = ttk.Frame(batch_frame)
        row1.pack(fill=tk.X, pady=5)
        
        ttk.Label(row1, text="注册数量:").pack(side=tk.LEFT, padx=(0, 5))
        self.batch_count_spinbox = ttk.Spinbox(row1, from_=1, to=100, width=8)
        self.batch_count_spinbox.set(10)
        self.batch_count_spinbox.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(row1, text="并发数:").pack(side=tk.LEFT, padx=(0, 5))
        self.batch_concurrent_spinbox = ttk.Spinbox(row1, from_=1, to=10, width=8)
        self.batch_concurrent_spinbox.set(3)
        self.batch_concurrent_spinbox.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(row1, text="密码模板:").pack(side=tk.LEFT, padx=(0, 5))
        self.batch_password_entry = ttk.Entry(row1, width=15)
        self.batch_password_entry.insert(0, "Aa123456")
        self.batch_password_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(row1, text="延迟(秒):").pack(side=tk.LEFT, padx=(0, 5))
        self.batch_delay_spinbox = ttk.Spinbox(row1, from_=3, to=30, width=8)
        self.batch_delay_spinbox.set(8)
        self.batch_delay_spinbox.pack(side=tk.LEFT)
        
        # 添加提示标签
        hint_label = ttk.Label(row1, text="💡 延迟越长越稳定", foreground='gray', font=('', 8))
        hint_label.pack(side=tk.LEFT, padx=5)
        
        # 第二行：控制按钮
        row2 = ttk.Frame(batch_frame)
        row2.pack(fill=tk.X, pady=5)
        
        self.batch_start_button = ttk.Button(row2, text="🚀 开始批量注册", 
                                             command=self.start_batch_register,
                                             style="Accent.TButton")
        self.batch_start_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.batch_pause_button = ttk.Button(row2, text="⏸️ 暂停", 
                                            command=self.pause_batch_register,
                                            state=tk.DISABLED)
        self.batch_pause_button.pack(side=tk.LEFT, padx=(0, 5))
        
        self.batch_stop_button = ttk.Button(row2, text="⏹️ 停止", 
                                           command=self.stop_batch_register,
                                           state=tk.DISABLED)
        self.batch_stop_button.pack(side=tk.LEFT)
        
        # 第三行：进度条
        row3 = ttk.Frame(batch_frame)
        row3.pack(fill=tk.X, pady=5)
        
        ttk.Label(row3, text="进度:").pack(side=tk.LEFT, padx=(0, 5))
        self.batch_progress = ttk.Progressbar(row3, mode='determinate', length=400)
        self.batch_progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.batch_status_label = ttk.Label(row3, text="0/0 成功:0 失败:0")
        self.batch_status_label.pack(side=tk.LEFT)
        
        # ========== 账号列表区域 ==========
        list_frame = ttk.LabelFrame(parent_frame, text="📋 账号列表", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 创建 Treeview 表格
        columns = ('id', 'email', 'password', 'session_id', 'did', 'create_time', 'status')
        self.account_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)
        
        # 设置列标题和宽度
        self.account_tree.heading('id', text='序号')
        self.account_tree.heading('email', text='邮箱')
        self.account_tree.heading('password', text='密码')
        self.account_tree.heading('session_id', text='Session ID')
        self.account_tree.heading('did', text='DID')
        self.account_tree.heading('create_time', text='创建时间')
        self.account_tree.heading('status', text='状态')
        
        self.account_tree.column('id', width=50, anchor='center')
        self.account_tree.column('email', width=180)
        self.account_tree.column('password', width=100)
        self.account_tree.column('session_id', width=200)
        self.account_tree.column('did', width=150)
        self.account_tree.column('create_time', width=150)
        self.account_tree.column('status', width=70, anchor='center')
        
        # 添加滚动条
        scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.account_tree.yview)
        scrollbar_x = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.account_tree.xview)
        self.account_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.account_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # 双击复制功能
        self.account_tree.bind('<Double-1>', self.on_account_double_click)
        
        # ========== 操作按钮区域 ==========
        action_frame = ttk.Frame(parent_frame)
        action_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Button(action_frame, text="📊 导出Excel", 
                  command=self.export_excel).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="📄 导出CSV", 
                  command=self.export_csv).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="🗑️ 删除选中", 
                  command=self.delete_selected_accounts).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="🧹 清空列表", 
                  command=self.clear_all_accounts).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="🔄 刷新", 
                  command=self.refresh_account_list).pack(side=tk.LEFT)
        
        # 统计信息
        self.account_stats_label = ttk.Label(action_frame, text="统计: 总计 0 个账号 | 可用 0 个 | 失效 0 个",
                                             foreground='gray')
        self.account_stats_label.pack(side=tk.RIGHT, padx=10)
        
        # 初始加载账号列表
        self.refresh_account_list()

    def _browse_file(self, entry_widget):
        filepath = filedialog.askopenfilename(title="选择图片文件", filetypes=[("Image Files", "*.png *.jpg *.jpeg *.webp")])
        if filepath:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filepath)
    
    def _toggle_end_frame(self):
        state = tk.NORMAL if self.use_first_last_var.get() else tk.DISABLED
        self.i2v_end_frame_label.config(state=state)
        self.i2v_end_frame_path.config(state=state)
        self.i2v_end_frame_browse_btn.config(state=state)

    def _toggle_login_method(self):
        """切换登录方式显示"""
        method = self.login_method_var.get()
        if method == "cookie":
            self.cookie_frame.pack(fill=tk.X)
            self.account_frame.pack_forget()
        else:  # password
            self.cookie_frame.pack_forget()
            self.account_frame.pack(fill=tk.X)
    
    def log(self, message):
        self.log_area.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {message}\n")
        self.log_area.see(tk.END)

    def choose_download_folder(self):
        """选择下载文件夹"""
        folder = filedialog.askdirectory(title="选择下载文件夹", initialdir=self.download_path)
        if folder:
            self.download_path = folder
            self.download_path_label.config(text=f"保存到: {self.download_path}")
            self.log(f"下载文件夹已设置为: {self.download_path}")
            # 自动保存配置
            self.save_config()
    
    def load_config(self):
        try:
            with open(self.CONFIG_FILE, 'r') as f:
                config = json.load(f)
                self.cookie_text.delete('1.0', tk.END); self.cookie_text.insert('1.0', config.get("cookies", ""))
                self.did_entry.delete(0, tk.END); self.did_entry.insert(0, config.get("did", ""))
                
                # 加载账号信息（如果有）
                if "email" in config:
                    self.email_entry.delete(0, tk.END)
                    self.email_entry.insert(0, config.get("email", ""))
                if "password" in config:
                    self.password_entry.delete(0, tk.END)
                    self.password_entry.insert(0, config.get("password", ""))
                
                # 加载登录方式
                if "login_method" in config:
                    self.login_method_var.set(config.get("login_method", "cookie"))
                    self._toggle_login_method()
                
                # 加载下载路径
                if "download_path" in config:
                    self.download_path = config["download_path"]
                    self.download_path_label.config(text=f"保存到: {self.download_path}")
                # 加载等待时间
                if "initial_wait_time" in config:
                    self.initial_wait_time = config["initial_wait_time"]
                    self.wait_time_spinbox.set(self.initial_wait_time)
            self.log("配置已加载。")
        except FileNotFoundError: self.log("未找到配置文件，请填写后保存。")
        except Exception as e: self.log(f"加载配置失败: {e}")

    def save_config(self):
        config = {
            "cookies": self.cookie_text.get('1.0', tk.END).strip(), 
            "did": self.did_entry.get().strip(),
            "login_method": self.login_method_var.get(),
            "download_path": self.download_path,
            "initial_wait_time": int(self.wait_time_spinbox.get())
        }
        
        # 可选：保存账号密码（用户自行决定是否保存）
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        if email:
            result = messagebox.askyesno("保存账号信息", 
                "是否保存账号和密码到配置文件？\n\n⚠️ 注意：密码将以明文形式保存在本地，请确保电脑安全。")
            if result:
                config["email"] = email
                config["password"] = password
        
        try:
            with open(self.CONFIG_FILE, 'w') as f: json.dump(config, f, indent=4)
            self.log("配置已保存到 capcut_config.json。")
        except Exception as e: self.log(f"保存配置失败: {e}")

    def _validate_config(self):
        cookies = self.cookie_text.get('1.0', tk.END).strip()
        did = self.did_entry.get().strip()
        if not cookies:
            messagebox.showerror("错误", "Cookie 或 sessionid 是必需的！")
            return None, None
        return cookies, did
    
    def start_credit_query(self):
        cookies, did = self._validate_config()
        if not cookies: return
        self.credit_button.config(state=tk.DISABLED, text="查询中...")
        thread = threading.Thread(target=self.generation_logic, args=("credit", cookies, did))
        thread.daemon = True; thread.start()

    def start_credit_receive(self):
        cookies, did = self._validate_config()
        if not cookies: return
        self.receive_credit_button.config(state=tk.DISABLED, text="领取中...")
        thread = threading.Thread(target=self.generation_logic, args=("credit_receive", cookies, did))
        thread.daemon = True; thread.start()

    def start_config_query(self):
        cookies, did = self._validate_config()
        if not cookies: return
        self.config_button.config(state=tk.DISABLED, text="查询中...")
        thread = threading.Thread(target=self.generation_logic, args=("config", cookies, did))
        thread.daemon = True; thread.start()

    def start_user_info_query(self):
        cookies, did = self._validate_config()
        if not cookies: return
        self.user_info_button.config(state=tk.DISABLED, text="查询中...")
        thread = threading.Thread(target=self.generation_logic, args=("user_info", cookies, did))
        thread.daemon = True; thread.start()

    def start_login(self):
        """开始登录流程"""
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        did = self.did_entry.get().strip()
        
        if not email or not password:
            messagebox.showerror("错误", "请输入邮箱和密码！")
            return
        
        self.login_button.config(state=tk.DISABLED, text="登录中...")
        thread = threading.Thread(target=self.login_logic, args=(email, password, did))
        thread.daemon = True
        thread.start()

    def start_t2i_generation(self):
        cookies, did = self._validate_config()
        if not cookies: return
        prompt = self.t2i_prompt_text.get('1.0', tk.END).strip()
        if not prompt: messagebox.showerror("错误", "请输入提示词！"); return
        neg_prompt = self.t2i_neg_prompt_text.get('1.0', tk.END).strip()
        
        # --- 已修改: 获取模型选择并传递 ---
        friendly_name = self.t2i_model_selector.get()
        model_key = self.MODEL_MAPPING.get(friendly_name, "high_aes_general_v30l:general_v3.0_18b") # 安全获取，失败则用默认
        should_download = self.download_images_var.get()
        wait_time = int(self.wait_time_spinbox.get())  # 获取等待时间
        
        self.t2i_button.config(state=tk.DISABLED, text="生成中...")
        
        thread = threading.Thread(target=self.generation_logic, args=("t2i", cookies, did, prompt, neg_prompt, model_key, should_download, wait_time))
        thread.daemon = True; thread.start()

    def start_t2v_generation(self):
        cookies, did = self._validate_config()
        if not cookies: return
        prompt = self.t2v_prompt_text.get('1.0', tk.END).strip()
        if not prompt: messagebox.showerror("错误", "请输入提示词！"); return
        ratio = self.video_ratio.get()
        duration_ms = int(self.duration_spinbox.get()) * 1000
        self.t2v_button.config(state=tk.DISABLED, text="生成中...")
        thread = threading.Thread(target=self.generation_logic, args=("t2v", cookies, did, prompt, ratio, duration_ms))
        thread.daemon = True; thread.start()

    def start_i2v_generation(self):
        cookies, did = self._validate_config()
        if not cookies: return

        # 获取选择的模式
        mode = self.i2v_mode_var.get()
        
        # 查询模式
        if mode == "query_only":
            submit_id = self.i2v_submit_id_entry.get().strip()
            if not submit_id:
                messagebox.showerror("错误", "请输入 Submit ID！")
                return
            self.i2v_button.config(state=tk.DISABLED, text="查询并下载中...")
            thread = threading.Thread(target=self.generation_logic, args=("i2v_query", cookies, did, submit_id))
            thread.daemon = True
            thread.start()
            return
        
        # 提交模式和完整流程模式
        prompt = self.i2v_prompt_text.get('1.0', tk.END).strip()
        start_frame_path = self.i2v_start_frame_path.get().strip()
        use_first_last = self.use_first_last_var.get()
        end_frame_path = self.i2v_end_frame_path.get().strip()
        ratio = self.i2v_video_ratio.get()
        duration_ms = int(self.i2v_duration_spinbox.get()) * 1000

        if not prompt: messagebox.showerror("错误", "请输入提示词！"); return
        if not start_frame_path: messagebox.showerror("错误", "请选择起始帧图片！"); return
        if use_first_last and not end_frame_path: messagebox.showerror("错误", "请选择结束帧图片！"); return

        if mode == "submit_only":
            self.i2v_button.config(state=tk.DISABLED, text="仅提交中...")
            thread = threading.Thread(target=self.generation_logic, args=("i2v_submit_only", cookies, did, prompt, start_frame_path, use_first_last, end_frame_path, ratio, duration_ms))
        else:  # mode == "full"
            self.i2v_button.config(state=tk.DISABLED, text="上传并生成中...")
            thread = threading.Thread(target=self.generation_logic, args=("i2v", cookies, did, prompt, start_frame_path, use_first_last, end_frame_path, ratio, duration_ms))
        
        thread.daemon = True
        thread.start()

    def login_logic(self, email, password, did):
        """登录逻辑处理（完整流程）"""
        try:
            self.api_queue.put(("log", "🚀 开始完整登录流程..."))
            self.api_queue.put(("log", f"📧 邮箱: {email}"))
            self.api_queue.put(("log", ""))
            
            # 创建临时 API 实例（不需要 cookies）
            api = CapCutAPI("", did)
            
            self.api_queue.put(("log", "步骤 1/4: 地区验证..."))
            self.api_queue.put(("log", "步骤 2/4: 邮箱密码登录..."))
            self.api_queue.put(("log", "步骤 3/4: 提交用户协议..."))
            self.api_queue.put(("log", "步骤 4/4: 获取用户详细信息..."))
            self.api_queue.put(("log", ""))
            
            result = api.login_with_email(email, password)
            
            if result.get("error"):
                self.api_queue.put(("log", f"❌ 登录失败: {result.get('error')}"))
                self.api_queue.put(("result", result))
            elif result.get('message') == 'success':
                sessionid = result.get('extracted_sessionid')
                
                if sessionid:
                    self.api_queue.put(("log", "=" * 50))
                    self.api_queue.put(("log", "✅ 登录成功！"))
                    self.api_queue.put(("log", "=" * 50))
                    self.api_queue.put(("log", f"\n🔑 Session ID: {sessionid}\n"))
                    
                    # 自动填充到 Cookie 输入框
                    self.cookie_text.delete('1.0', tk.END)
                    self.cookie_text.insert('1.0', sessionid)
                    
                    self.api_queue.put(("log", "✅ Session ID 已自动填充到配置区域"))
                    self.api_queue.put(("log", "💡 提示: 可以点击 '保存配置' 按钮保存"))
                    
                    # 显示基本用户信息
                    user_data = result.get('data', {})
                    if 'user_id' in user_data:
                        self.api_queue.put(("log", "\n" + "=" * 50))
                        self.api_queue.put(("log", "📋 基本账号信息"))
                        self.api_queue.put(("log", "=" * 50))
                        self.api_queue.put(("log", f"👤 用户名: {user_data.get('name', 'N/A')}"))
                        self.api_queue.put(("log", f"🆔 用户ID: {user_data.get('user_id_str', 'N/A')}"))
                        self.api_queue.put(("log", f"📧 邮箱: {user_data.get('email', 'N/A')}"))
                        self.api_queue.put(("log", f"🌍 地区: {user_data.get('store_country', 'N/A')}"))
                    
                    # 显示详细用户信息
                    detailed_info = result.get('detailed_user_info', {})
                    if detailed_info:
                        user_info = detailed_info.get('user_info', {})
                        space_info = detailed_info.get('space_info', {})
                        location = detailed_info.get('location', {})
                        
                        if user_info:
                            self.api_queue.put(("log", "\n" + "=" * 50))
                            self.api_queue.put(("log", "🎨 详细账号信息"))
                            self.api_queue.put(("log", "=" * 50))
                            
                            # 判断是否新用户
                            if user_info.get('is_new_user'):
                                self.api_queue.put(("log", "🎉 欢迎新用户！"))
                            
                            create_time = user_info.get('create_time', 0)
                            if create_time:
                                import time as time_module
                                create_date = time_module.strftime('%Y-%m-%d', time_module.localtime(create_time))
                                self.api_queue.put(("log", f"📅 注册日期: {create_date}"))
                            
                            self.api_queue.put(("log", f"✉️ 绑定邮箱: {user_info.get('bind_email', 'N/A')}"))
                            
                            # TikTok绑定状态
                            if user_info.get('is_bind_tt'):
                                self.api_queue.put(("log", "🔗 已绑定 TikTok 账号"))
                        
                        if space_info:
                            self.api_queue.put(("log", "\n" + "=" * 50))
                            self.api_queue.put(("log", "💾 工作空间信息"))
                            self.api_queue.put(("log", "=" * 50))
                            self.api_queue.put(("log", f"🏢 空间ID: {space_info.get('space_id', 'N/A')}"))
                            self.api_queue.put(("log", f"🌐 空间区域: {space_info.get('space_idc', 'N/A')}"))
                            self.api_queue.put(("log", f"📂 工作区ID: {space_info.get('workspace_id', 'N/A')}"))
                        
                        if location:
                            domain_info = location.get('domain', {})
                            if domain_info:
                                self.api_queue.put(("log", "\n" + "=" * 50))
                                self.api_queue.put(("log", "🔗 服务域名信息"))
                                self.api_queue.put(("log", "=" * 50))
                                self.api_queue.put(("log", f"编辑服务: {domain_info.get('web_domain', 'N/A')}"))
                                self.api_queue.put(("log", f"商业服务: {domain_info.get('commerce_domain', 'N/A')}"))
                    
                    self.api_queue.put(("log", "\n" + "=" * 50))
                    self.api_queue.put(("log", "🎊 登录流程完成！现在可以使用所有功能了"))
                    self.api_queue.put(("log", "=" * 50))
                else:
                    self.api_queue.put(("log", "⚠️ 登录响应成功，但未能自动提取 sessionid"))
                    self.api_queue.put(("log", "请查看下方完整响应数据，手动提取 sessionid"))
                
                if result.get('detailed_user_info'):
                    # 不显示详细的完整响应，避免太长
                    self.api_queue.put(("log", "\n💡 详细信息已解析显示，完整数据已省略"))
                else:
                    self.api_queue.put(("log", "\n完整登录响应:"))
                    self.api_queue.put(("result", result))
            else:
                self.api_queue.put(("log", "❌ 登录失败"))
                self.api_queue.put(("result", result))
                
                # 检查常见错误
                error_code = result.get('error_code', 0)
                desc = result.get('description', '')
                if error_code == 1006:
                    self.api_queue.put(("log", "\n💡 提示: 密码错误，请检查您的账号密码"))
                elif error_code == 1009:
                    self.api_queue.put(("log", "\n💡 提示: 需要验证码，请使用浏览器登录后获取 Cookie"))
                elif desc:
                    self.api_queue.put(("log", f"\n💡 错误信息: {desc}"))
                    
        except Exception as e:
            self.api_queue.put(("log", f"❌ 登录过程发生错误: {e}"))
            import traceback
            self.api_queue.put(("log", traceback.format_exc()))
        finally:
            self.api_queue.put(("done", "login"))

    def start_register(self):
        """开始注册流程"""
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        did = self.did_entry.get().strip()
        
        if not email or not password:
            messagebox.showerror("错误", "请输入邮箱和密码！")
            return
        
        # 验证邮箱格式
        import re
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            messagebox.showerror("错误", "请输入有效的邮箱地址！")
            return
        
        self.register_button.config(state=tk.DISABLED, text="注册中...")
        thread = threading.Thread(target=self.register_logic, args=(email, password, did))
        thread.daemon = True
        thread.start()
    
    def start_auto_register(self):
        """开始全自动注册流程（自动生成临时邮箱+自动获取验证码）"""
        password = self.password_entry.get().strip()
        did = self.did_entry.get().strip()
        
        # 如果没有填写密码，使用默认密码
        if not password:
            password = "Aa123456"  # 默认密码
            self.password_entry.delete(0, tk.END)
            self.password_entry.insert(0, password)
        
        # 确认开始
        result = messagebox.askyesno(
            "🤖 全自动注册", 
            f"将使用以下设置自动注册新账号：\n\n"
            f"📧 邮箱: 自动生成临时邮箱\n"
            f"🔐 密码: {password}\n"
            f"🎂 生日: 2000-01-01\n\n"
            f"整个过程完全自动化，无需任何操作！\n"
            f"是否继续？"
        )
        
        if not result:
            return
        
        self.auto_register_button.config(state=tk.DISABLED, text="全自动注册中...")
        thread = threading.Thread(target=self.auto_register_logic, args=(password, did))
        thread.daemon = True
        thread.start()
    
    def auto_register_logic(self, password, did):
        """全自动注册逻辑处理（自动生成邮箱+自动获取验证码）"""
        try:
            import hashlib
            import time
            
            self.api_queue.put(("log", "🤖 ============== 全自动注册开始 =============="))
            self.api_queue.put(("log", ""))
            
            # 创建临时 API 实例
            api = CapCutAPI("", did)
            
            # 步骤1: 生成临时邮箱
            self.api_queue.put(("log", "📧 步骤 1/5: 生成临时邮箱..."))
            email = api.get_temp_email()
            self.api_queue.put(("log", f"✅ 临时邮箱: {email}"))
            self.api_queue.put(("log", ""))
            
            # 自动填充邮箱到输入框
            self.after(0, lambda: self.email_entry.delete(0, tk.END))
            self.after(0, lambda: self.email_entry.insert(0, email))
            
            # 步骤2: 发送验证码
            self.api_queue.put(("log", "📨 步骤 2/5: 发送验证码到临时邮箱..."))
            send_result = api.send_register_code(email, password)
            
            if not send_result.get('success'):
                self.api_queue.put(("log", f"❌ 发送验证码失败: {send_result.get('error', '未知错误')}"))
                if send_result.get('details'):
                    self.api_queue.put(("result", send_result['details']))
                self.api_queue.put(("done", "auto_register"))
                return
            
            self.api_queue.put(("log", f"✅ {send_result.get('message', '验证码已发送')}"))
            self.api_queue.put(("log", ""))
            
            # 步骤3: 自动获取验证码
            self.api_queue.put(("log", "🔍 步骤 3/5: 自动监控邮箱获取验证码..."))
            self.api_queue.put(("log", "⏳ 正在轮询邮箱，请稍候（最多等待120秒）..."))
            
            code_result = api.check_verification_code(email, timeout=120)
            
            if not code_result.get('success'):
                self.api_queue.put(("log", f"❌ {code_result.get('error', '获取验证码失败')}"))
                self.api_queue.put(("done", "auto_register"))
                return
            
            code = code_result.get('code')
            self.api_queue.put(("log", f"✅ 成功获取验证码: {code}"))
            self.api_queue.put(("log", ""))
            
            # 步骤4: 验证验证码
            self.api_queue.put(("log", "✨ 步骤 4/5: 验证验证码..."))
            
            verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
            verify_url = "https://dreamina.capcut.com/passport/web/email/register/code_verify/"
            verify_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            encoded_email = api._encode_login_data(email)
            encoded_code = api._encode_login_data(code)
            
            from urllib.parse import urlencode
            verify_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,
                'code': encoded_code,
                'type': '34',
                'fixed_mix_mode': '1'
            })
            
            csrf_token = None
            for cookie in api.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            verify_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': api.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home',
                'user-agent': api.user_agent,
            }
            
            if csrf_token:
                verify_headers['x-tt-passport-csrf-token'] = csrf_token
            
            try:
                verify_response = api.session.post(
                    verify_url,
                    params=verify_params,
                    headers=verify_headers,
                    data=verify_data,
                    timeout=15
                )
                verify_response.raise_for_status()
                verify_result = verify_response.json()
                
                if verify_result.get('message') != 'success':
                    self.api_queue.put(("log", f"❌ 验证码验证失败: {verify_result.get('data', {}).get('description', '未知错误')}"))
                    self.api_queue.put(("result", verify_result))
                    self.api_queue.put(("done", "auto_register"))
                    return
                
                # 获取新的 email_ticket
                new_email_ticket = verify_result.get('data', {}).get('email_ticket', '')
                self.api_queue.put(("log", "✅ 验证码验证成功"))
                self.api_queue.put(("log", ""))
                
            except Exception as e:
                self.api_queue.put(("log", f"❌ 验证验证码时出错: {e}"))
                self.api_queue.put(("done", "auto_register"))
                return
            
            # 步骤5: 完成注册
            birthday = "2000-01-01"
            self.api_queue.put(("log", "🎉 步骤 5/5: 提交注册信息并完成注册..."))
            register_result = api.register_with_email(email, password, code, new_email_ticket, birthday)
            
            if register_result.get('success'):
                self.api_queue.put(("log", "\n" + "=" * 50))
                self.api_queue.put(("log", "🎊 全自动注册成功！"))
                self.api_queue.put(("log", "=" * 50))
                
                # 显示账号信息
                session_id = register_result.get('session_id')
                user_info = register_result.get('user_info', {})
                
                if session_id:
                    self.api_queue.put(("log", f"\n🔑 Session ID: {session_id}"))
                    self.api_queue.put(("log", "✅ Session 已自动提取"))
                
                self.api_queue.put(("log", "\n" + "=" * 50))
                self.api_queue.put(("log", "📋 新账号信息"))
                self.api_queue.put(("log", "=" * 50))
                self.api_queue.put(("log", f"📧 邮箱: {email}"))
                self.api_queue.put(("log", f"🔐 密码: {password}"))
                self.api_queue.put(("log", f"🎂 生日: {birthday}"))
                
                if user_info:
                    self.api_queue.put(("log", f"👤 用户名: {user_info.get('screen_name', 'N/A')}"))
                    self.api_queue.put(("log", f"🆔 用户ID: {user_info.get('user_id_str', 'N/A')}"))
                    self.api_queue.put(("log", f"🌍 地区: {user_info.get('store_country', 'N/A')}"))
                
                # 自动填充 Cookie
                if session_id:
                    self.after(0, lambda: self.cookie_text.delete('1.0', tk.END))
                    self.after(0, lambda: self.cookie_text.insert('1.0', f"sessionid={session_id}"))
                    self.api_queue.put(("log", "\n💡 Session 已自动填充到配置中"))
                    self.api_queue.put(("log", "💡 建议现在点击【保存配置】保存账号信息"))
                
                self.api_queue.put(("log", "\n" + "=" * 50))
                self.api_queue.put(("log", "🎉 完成！现在可以立即使用所有功能了"))
                self.api_queue.put(("log", "=" * 50))
            else:
                self.api_queue.put(("log", f"\n❌ 注册失败: {register_result.get('error', '未知错误')}"))
                if register_result.get('details'):
                    self.api_queue.put(("result", register_result['details']))
                
        except Exception as e:
            self.api_queue.put(("log", f"❌ 全自动注册过程发生错误: {e}"))
            import traceback
            self.api_queue.put(("log", traceback.format_exc()))
        finally:
            self.api_queue.put(("done", "auto_register"))
    
    def register_logic(self, email, password, did):
        """注册逻辑处理"""
        try:
            import hashlib
            import time
            verify_fp = f"verify_{hashlib.md5(str(time.time()).encode()).hexdigest()[:16]}"
            
            self.api_queue.put(("log", "🎉 开始注册流程..."))
            self.api_queue.put(("log", f"📧 邮箱: {email}"))
            self.api_queue.put(("log", ""))
            
            # 创建临时 API 实例
            api = CapCutAPI("", did)
            
            # 步骤1: 发送验证码
            self.api_queue.put(("log", "📨 步骤 1/3: 发送验证码到邮箱..."))
            send_result = api.send_register_code(email, password)
            
            if not send_result.get('success'):
                self.api_queue.put(("log", f"❌ 发送验证码失败: {send_result.get('error', '未知错误')}"))
                if send_result.get('details'):
                    self.api_queue.put(("result", send_result['details']))
                self.api_queue.put(("done", "register"))
                return
            
            self.api_queue.put(("log", f"✅ {send_result.get('message', '验证码已发送')}"))
            self.api_queue.put(("log", ""))
            
            # 步骤2: 等待用户输入验证码
            self.api_queue.put(("log", "📬 请查收邮件中的验证码..."))
            self.api_queue.put(("log", ""))
            
            # 保存 email_ticket
            email_ticket = send_result.get('email_ticket', '')
            
            # 使用共享变量和事件来同步线程
            import threading
            input_event = threading.Event()
            input_result = {'code': None, 'birthday': None}
            
            def ask_info_in_main_thread():
                """在主线程中显示对话框"""
                from tkinter import simpledialog
                import tkinter as tk
                from tkinter import ttk
                
                # 创建自定义对话框
                dialog = tk.Toplevel(self)
                dialog.title("完成注册")
                dialog.geometry("400x200")
                dialog.transient(self)
                dialog.grab_set()
                
                # 居中显示
                dialog.update_idletasks()
                x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
                y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
                dialog.geometry(f"+{x}+{y}")
                
                frame = ttk.Frame(dialog, padding="20")
                frame.pack(fill=tk.BOTH, expand=True)
                
                # 提示信息
                ttk.Label(frame, text=f"验证码已发送到 {send_result.get('email', email)}", 
                         font=('', 10, 'bold')).pack(pady=(0, 15))
                
                # 验证码输入
                code_frame = ttk.Frame(frame)
                code_frame.pack(fill=tk.X, pady=5)
                ttk.Label(code_frame, text="验证码:", width=10).pack(side=tk.LEFT)
                code_entry = ttk.Entry(code_frame)
                code_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                code_entry.focus()
                
                # 生日输入
                birthday_frame = ttk.Frame(frame)
                birthday_frame.pack(fill=tk.X, pady=5)
                ttk.Label(birthday_frame, text="生日:", width=10).pack(side=tk.LEFT)
                birthday_entry = ttk.Entry(birthday_frame)
                birthday_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
                birthday_entry.insert(0, "2000-01-01")
                
                ttk.Label(frame, text="(格式: YYYY-MM-DD)", 
                         foreground='gray').pack(pady=(0, 10))
                
                # 按钮
                def on_confirm():
                    input_result['code'] = code_entry.get().strip()
                    input_result['birthday'] = birthday_entry.get().strip()
                    dialog.destroy()
                    input_event.set()
                
                def on_cancel():
                    input_result['code'] = None
                    input_result['birthday'] = None
                    dialog.destroy()
                    input_event.set()
                
                btn_frame = ttk.Frame(frame)
                btn_frame.pack(pady=10)
                ttk.Button(btn_frame, text="确定", command=on_confirm, 
                          style="Accent.TButton").pack(side=tk.LEFT, padx=5)
                ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT)
                
                # 绑定回车键
                code_entry.bind('<Return>', lambda e: on_confirm())
                birthday_entry.bind('<Return>', lambda e: on_confirm())
                
                dialog.protocol("WM_DELETE_WINDOW", on_cancel)
            
            # 在主线程中显示对话框
            self.after(0, ask_info_in_main_thread)
            
            # 等待用户输入（最多2分钟）
            self.api_queue.put(("log", "⏳ 等待输入验证码和生日..."))
            if not input_event.wait(timeout=120):
                self.api_queue.put(("log", "⏱️ 输入超时"))
                self.api_queue.put(("done", "register"))
                return
            
            code = input_result['code']
            birthday = input_result['birthday']
            
            if not code or not birthday:
                self.api_queue.put(("log", "❌ 取消输入或信息不完整"))
                self.api_queue.put(("done", "register"))
                return
            
            self.api_queue.put(("log", f"🔑 验证码: {code}"))
            self.api_queue.put(("log", ""))
            
            # 步骤3: 先验证验证码，获取新的 email_ticket
            self.api_queue.put(("log", "✨ 步骤 2/3: 验证验证码..."))
            
            verify_url = "https://dreamina.capcut.com/passport/web/email/register/code_verify/"
            verify_params = {
                'aid': '513641',
                'account_sdk_source': 'web',
                'sdk_version': '2.1.10-tiktok',
                'language': 'en',
                'verifyFp': verify_fp
            }
            
            encoded_email = api._encode_login_data(email)
            encoded_code = api._encode_login_data(code)
            
            from urllib.parse import urlencode
            verify_data = urlencode({
                'mix_mode': '1',
                'email': encoded_email,
                'code': encoded_code,
                'type': '34',
                'fixed_mix_mode': '1'
            })
            
            csrf_token = None
            for cookie in api.session.cookies:
                if cookie.name == 'passport_csrf_token':
                    csrf_token = cookie.value
                    break
            
            verify_headers = {
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/x-www-form-urlencoded',
                'did': api.did,
                'appid': '513641',
                'origin': 'https://dreamina.capcut.com',
                'referer': 'https://dreamina.capcut.com/ai-tool/home',
                'user-agent': api.user_agent,
            }
            
            if csrf_token:
                verify_headers['x-tt-passport-csrf-token'] = csrf_token
            
            try:
                verify_response = api.session.post(
                    verify_url,
                    params=verify_params,
                    headers=verify_headers,
                    data=verify_data,
                    timeout=15
                )
                verify_response.raise_for_status()
                verify_result = verify_response.json()
                
                if verify_result.get('message') != 'success':
                    self.api_queue.put(("log", f"❌ 验证码验证失败: {verify_result.get('data', {}).get('description', '未知错误')}"))
                    self.api_queue.put(("result", verify_result))
                    self.api_queue.put(("done", "register"))
                    return
                
                # 获取新的 email_ticket
                new_email_ticket = verify_result.get('data', {}).get('email_ticket', '')
                self.api_queue.put(("log", "✅ 验证码验证成功"))
                self.api_queue.put(("log", ""))
                
            except Exception as e:
                self.api_queue.put(("log", f"❌ 验证验证码时出错: {e}"))
                self.api_queue.put(("done", "register"))
                return
            
            # 步骤4: 输入生日完成注册
            self.api_queue.put(("log", f"🎂 生日: {birthday}"))
            self.api_queue.put(("log", "✨ 步骤 3/3: 提交生日并完成注册..."))
            register_result = api.register_with_email(email, password, code, new_email_ticket, birthday)
            
            if register_result.get('success'):
                self.api_queue.put(("log", "\n" + "=" * 50))
                self.api_queue.put(("log", "🎊 注册成功！"))
                self.api_queue.put(("log", "=" * 50))
                
                # 显示用户信息
                session_id = register_result.get('session_id')
                user_info = register_result.get('user_info', {})
                
                if session_id:
                    self.api_queue.put(("log", f"\n🔑 Session ID: {session_id}"))
                    self.api_queue.put(("log", "✅ Session 已自动提取，可以保存配置使用"))
                
                if user_info:
                    self.api_queue.put(("log", "\n" + "=" * 50))
                    self.api_queue.put(("log", "👤 账号信息"))
                    self.api_queue.put(("log", "=" * 50))
                    self.api_queue.put(("log", f"👤 用户名: {user_info.get('screen_name', 'N/A')}"))
                    self.api_queue.put(("log", f"🆔 用户ID: {user_info.get('user_id_str', 'N/A')}"))
                    self.api_queue.put(("log", f"📧 邮箱: {user_info.get('email', 'N/A')}"))
                    self.api_queue.put(("log", f"🌍 地区: {user_info.get('store_country', 'N/A')}"))
                
                # 自动填充 Cookie
                if session_id:
                    self.after(0, lambda: self.cookie_text.delete('1.0', tk.END))
                    self.after(0, lambda: self.cookie_text.insert('1.0', f"sessionid={session_id}"))
                    self.api_queue.put(("log", "\n💡 Session 已自动填充到配置中，记得点击【保存配置】"))
                
                self.api_queue.put(("log", "\n" + "=" * 50))
                self.api_queue.put(("log", "🎉 注册完成！现在可以使用所有功能了"))
                self.api_queue.put(("log", "=" * 50))
            else:
                self.api_queue.put(("log", f"\n❌ 注册失败: {register_result.get('error', '未知错误')}"))
                if register_result.get('details'):
                    self.api_queue.put(("result", register_result['details']))
                
        except Exception as e:
            self.api_queue.put(("log", f"❌ 注册过程发生错误: {e}"))
            import traceback
            self.api_queue.put(("log", traceback.format_exc()))
        finally:
            self.api_queue.put(("done", "register"))

    def generation_logic(self, task_type, cookies, did, *args):
        try:
            self.api_queue.put(("log", "正在初始化API客户端..."))
            api = CapCutAPI(cookies, did)

            if task_type == "credit":
                self.api_queue.put(("log", "正在查询用户积分..."))
                result = api.get_user_credit()
                self.api_queue.put(("log", "查询完成！服务器返回:"))
                self.api_queue.put(("result", result))

            elif task_type == "credit_receive":
                self.api_queue.put(("log", "正在领取每日积分..."))
                result = api.credit_receive()
                
                # 尝试解析并美化积分领取结果
                try:
                    if result.get("ret") == "0" and "data" in result:
                        data = result.get("data", {})
                        receive_quota = data.get("receive_quota", 0)
                        is_first = data.get("is_first_receive", False)
                        
                        self.api_queue.put(("log", "\n✅ 积分领取成功！\n"))
                        self.api_queue.put(("log", f"🎁 本次领取: {receive_quota} 积分"))
                        
                        if is_first:
                            self.api_queue.put(("log", "🎉 这是您的首次领取！"))
                        else:
                            self.api_queue.put(("log", "📅 今日签到成功"))
                        
                        self.api_queue.put(("log", "\n💡 提示: 可以点击 '💰 查询积分' 查看当前积分余额"))
                        self.api_queue.put(("log", ""))
                    elif result.get("error"):
                        self.api_queue.put(("log", f"❌ 领取失败: {result.get('error')}"))
                    else:
                        # 检查是否已经领取过
                        errmsg = result.get("errmsg", "")
                        if "already" in errmsg.lower() or "领取" in errmsg:
                            self.api_queue.put(("log", "ℹ️ 今日积分已领取过了"))
                        else:
                            self.api_queue.put(("log", f"⚠️ 领取结果未知: {errmsg}"))
                except Exception as e:
                    self.api_queue.put(("log", f"解析领取结果时出错: {e}"))
                
                self.api_queue.put(("log", "\n完整服务器响应:"))
                self.api_queue.put(("result", result))

            elif task_type == "config":
                self.api_queue.put(("log", "正在获取通用配置信息..."))
                result = api.get_common_config()
                self.api_queue.put(("log", "获取完成！服务器返回:"))
                
                # 尝试解析并美化模型列表
                try:
                    if result.get("ret") == "0" and "data" in result:
                        model_list = result.get("data", {}).get("model_list", [])
                        self.api_queue.put(("log", f"\n✅ 成功获取 {len(model_list)} 个可用模型:\n"))
                        for idx, model in enumerate(model_list, 1):
                            model_name = model.get("model_name", "Unknown")
                            model_key = model.get("model_req_key", "Unknown")
                            model_tip = model.get("model_tip", "")
                            is_new = "🆕" if model.get("is_new_model", False) else ""
                            self.api_queue.put(("log", f"{idx}. {is_new}{model_name}"))
                            self.api_queue.put(("log", f"   模型密钥: {model_key}"))
                            if model_tip:
                                self.api_queue.put(("log", f"   描述: {model_tip}"))
                            self.api_queue.put(("log", ""))
                        self.api_queue.put(("log", "\n完整原始数据:"))
                except Exception as e:
                    self.api_queue.put(("log", f"解析模型列表时出错: {e}"))
                
                self.api_queue.put(("result", result))

            elif task_type == "user_info":
                self.api_queue.put(("log", "正在获取用户信息..."))
                result = api.get_user_info_region()
                self.api_queue.put(("log", "获取完成！"))
                
                # 尝试解析并美化用户信息
                try:
                    if result.get("message") == "success" and "data" in result:
                        user_data = result.get("data", {})
                        self.api_queue.put(("log", "\n✅ 用户信息获取成功:\n"))
                        
                        # 显示关键用户信息
                        if "user_id" in user_data:
                            self.api_queue.put(("log", f"👤 用户名: {user_data.get('name', 'N/A')}"))
                            self.api_queue.put(("log", f"🆔 用户ID: {user_data.get('user_id_str', user_data.get('user_id', 'N/A'))}"))
                            self.api_queue.put(("log", f"📧 邮箱: {user_data.get('email', 'N/A')}"))
                            self.api_queue.put(("log", f"🌍 地区: {user_data.get('store_country', 'N/A')}"))
                            self.api_queue.put(("log", f"🔐 SecUserID: {user_data.get('sec_user_id', 'N/A')}"))
                            self.api_queue.put(("log", f"📅 创建时间: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(user_data.get('user_create_time', 0)))}"))
                            self.api_queue.put(("log", ""))
                        elif "country_code" in user_data:
                            # 只有地区信息
                            self.api_queue.put(("log", f"🌍 国家/地区代码: {user_data.get('country_code', 'N/A')}"))
                            self.api_queue.put(("log", f"🔗 域名: {user_data.get('domain', 'N/A')}"))
                            self.api_queue.put(("log", ""))
                        
                        self.api_queue.put(("log", "\n完整原始数据:"))
                except Exception as e:
                    self.api_queue.put(("log", f"解析用户信息时出错: {e}"))
                
                self.api_queue.put(("result", result))

            elif task_type == "t2i":
                # --- 已修改: 接收 model_key 和 wait_time ---
                prompt, neg_prompt, model_key, should_download, wait_time = args
                self.api_queue.put(("log", f"正在使用模型 [{model_key}] 提交文生图任务..."))
                
                # 步骤1: 提交任务并获取 submit_id
                creation_response = api.create_text_to_image_task(prompt, neg_prompt, model_key)
                submit_id = None
                try:
                    submit_id = creation_response.get("data", {}).get("aigc_data", {}).get("submit_id")
                except Exception:
                    pass

                if not submit_id:
                    self.api_queue.put(("log", f"错误：任务提交失败或未能获取submit_id。"))
                    self.api_queue.put(("result", creation_response))
                    return

                self.api_queue.put(("log", f"任务提交成功！Submit ID: {submit_id}"))
                self.api_queue.put(("log", f"⏱️ 等待 {wait_time} 秒后开始查询任务状态..."))
                time.sleep(wait_time)  # 先等待指定时间
                self.api_queue.put(("log", "开始轮询任务状态..."))

                # 步骤2: 轮询查询任务状态
                max_retries = 15
                query_response = None
                for i in range(max_retries):
                    if i > 0:
                        time.sleep(3)  # 后续查询等待3秒
                    
                    self.api_queue.put(("log", f"正在进行第 {i + 1}/{max_retries} 次查询..."))
                    query_response = api.query_task_status(submit_id)
                    
                    task_data = query_response.get("data", {}).get(submit_id, {})
                    if not task_data:
                        self.api_queue.put(("log", f"警告：查询响应中未找到ID [{submit_id}] 的数据，继续..."))
                        continue
                    
                    status = task_data.get("status")
                    item_list = task_data.get("item_list", [])
                    self.api_queue.put(("log", f"  查询结果: status={status}, item_list中有{len(item_list)}个项目"))

                    if status != 20:
                        self.api_queue.put(("log", f"任务已结束，最终状态码: {status}"))
                        break
                
                # 步骤3: 处理最终结果
                try:
                    final_task_data = query_response.get("data", {}).get(submit_id, {})
                    final_item_list = final_task_data.get("item_list", [])

                    if not final_item_list:
                        self.api_queue.put(("log", "任务完成，但最终未能获取到任何图片项目。"))
                        self.api_queue.put(("result", query_response))
                        return

                    self.api_queue.put(("log", f"🎨 图片生成成功！共找到 {len(final_item_list)} 张图片。"))

                    if should_download:
                        download_path = self.download_path
                        os.makedirs(download_path, exist_ok=True)
                        self.api_queue.put(("log", f"图片将保存到: {os.path.abspath(download_path)}"))

                    for i, item in enumerate(final_item_list):
                        high_res_url = None
                        try:
                            high_res_url = item['image']['large_images'][0]['image_url']
                        except (KeyError, IndexError, TypeError):
                            pass
                        
                        image_url = high_res_url or item.get('common_attr', {}).get('cover_url')

                        if not image_url:
                            self.api_queue.put(("log", f"  - 第 {i+1} 张图片未找到URL，跳过。"))
                            continue
                        
                        self.api_queue.put(("log", f"============== 图片 {i+1} 链接 =============="))
                        self.api_queue.put(("log", image_url))
                        self.api_queue.put(("log", "======================================"))

                        if should_download:
                            filename = f"{time.strftime('%Y%m%d_%H%M%S')}_{i+1}.jpeg"
                            filepath = os.path.join(download_path, filename)
                            self.api_queue.put(("log", f"  - 正在下载第 {i+1} 张图片到 {filepath} ..."))
                            try:
                                with requests.get(image_url, stream=True, timeout=60) as r:
                                    r.raise_for_status()
                                    with open(filepath, 'wb') as f:
                                        for chunk in r.iter_content(chunk_size=8192):
                                            f.write(chunk)
                                self.api_queue.put(("log", f"  ✅ 图片 {i+1} 下载成功！"))
                            except Exception as e:
                                self.api_queue.put(("log", f"  ❌ 下载图片 {i+1} 失败: {e}"))
                                
                except Exception as parse_error:
                    self.api_queue.put(("log", f"解析或下载图片时出错: {parse_error}"))
                    import traceback
                    self.api_queue.put(("log", traceback.format_exc()))
                    self.api_queue.put(("result", query_response))
            
            elif task_type == "i2v_submit_only":
                # 新增：仅提交模式（不轮询、不下载）
                prompt, start_path, use_first, end_path, ratio, duration_ms = args
                self.api_queue.put(("log", "开始处理图生视频任务（仅提交模式）..."))
                
                def _upload_image(image_path, log_name):
                    self.api_queue.put(("log", f"正在读取 {log_name} 图片: {image_path}"))
                    with open(image_path, 'rb') as f: image_bytes = f.read()
                    
                    self.api_queue.put(("log", f"1/4: 为 {log_name} 获取上传Token..."))
                    token_res = api.get_upload_token()
                    if token_res.get("error"): raise Exception(f"获取Token失败: {token_res.get('details')}")
                    
                    self.api_queue.put(("log", f"2/4: 为 {log_name} 申请上传..."))
                    apply_res = api.apply_image_upload(len(image_bytes), token_res["data"])
                    if apply_res.get("error"): raise Exception(f"申请上传失败: {apply_res.get('details')}")
                    
                    upload_info = apply_res["Result"]["UploadAddress"]["StoreInfos"][0]
                    upload_info['UploadHost'] = apply_res["Result"]["UploadAddress"]["UploadHosts"][0]
                    session_key = apply_res["Result"]["UploadAddress"]["SessionKey"]
                    
                    self.api_queue.put(("log", f"3/4: 上传 {log_name} 文件..."))
                    upload_res = api.upload_image_file(image_bytes, upload_info)
                    if upload_res.get("error"): raise Exception(f"上传文件失败: {upload_res.get('error')}")
                    
                    self.api_queue.put(("log", f"4/4: 确认 {log_name} 上传..."))
                    commit_res = api.commit_image_upload(session_key, token_res["data"])
                    if commit_res.get("error"): raise Exception(f"确认上传失败: {commit_res.get('details')}")
                    
                    image_uri = commit_res["Result"]["PluginResult"][0]["ImageUri"]
                    self.api_queue.put(("log", f"{log_name} 上传成功! URI: {image_uri}"))
                    return image_uri

                start_uri = _upload_image(start_path, "起始帧")
                end_uri = _upload_image(end_path, "结束帧") if use_first else None
                
                self.api_queue.put(("log", "图片上传完成，正在提交视频生成任务..."))
                creation_response = api.create_image_to_video_task(prompt, start_uri, ratio, "720p", duration_ms, "Video 3.0", use_first, end_uri)
                submit_id = creation_response.get("data", {}).get("aigc_data", {}).get("submit_id")
                
                if not submit_id:
                    self.api_queue.put(("log", f"错误：任务提交失败或未能获取submit_id。"))
                    self.api_queue.put(("result", creation_response))
                else:
                    self.api_queue.put(("log", f"✅ 任务提交成功！"))
                    self.api_queue.put(("log", "============== Submit ID =============="))
                    self.api_queue.put(("log", submit_id))
                    self.api_queue.put(("log", "======================================"))
                    self.api_queue.put(("log", f"📋 请保存此 Submit ID，稍后可用于查询和下载视频"))
                return
            
            elif task_type == "i2v_query":
                # 新增：查询模式（输入submit_id）
                submit_id = args[0]
                self.api_queue.put(("log", f"开始查询任务: {submit_id}"))
                self.api_queue.put(("log", "开始轮询任务状态... (请耐心等待，可能需几分钟)"))
                
                max_retries = 30
                for i in range(max_retries):
                    time.sleep(10)
                    self.api_queue.put(("log", f"正在进行第 {i + 1}/{max_retries} 次查询..."))
                    query_response = api.query_video_task_status(submit_id)
                    
                    task_data = query_response.get("data", {}).get(submit_id, {})
                    if not task_data:
                        self.api_queue.put(("log", f"警告：查询响应中未找到ID [{submit_id}] 的数据，继续..."))
                        continue
                    
                    status = task_data.get("status")
                    item_list = task_data.get("item_list", [])
                    self.api_queue.put(("log", f"  查询结果: status={status}, item_list中有{len(item_list)}个项目"))
                    
                    if status != 20:
                        self.api_queue.put(("log", f"任务已结束，最终状态码: {status}"))
                        
                        video_url = None
                        try:
                            if item_list and isinstance(item_list, list) and len(item_list) > 0:
                                item = item_list[0]
                                video_info = item.get("video", {})
                                transcoded_video = video_info.get("transcoded_video", {})
                                origin_video = transcoded_video.get("origin", {})
                                video_url = origin_video.get("video_url")
                        except (AttributeError, IndexError, TypeError) as e:
                            self.api_queue.put(("log", f"解析视频链接时遇到错误: {e}, 将打印原始数据。"))
                        
                        if video_url:
                            self.api_queue.put(("log", "🎉 视频生成成功！"))
                            self.api_queue.put(("log", "============== 视频链接 =============="))
                            self.api_queue.put(("log", video_url))
                            self.api_queue.put(("log", "======================================"))
                        else:
                            self.api_queue.put(("log", f"任务已结束但未能从复杂结构中找到视频链接。"))
                            self.api_queue.put(("result", query_response))
                        return
                
                self.api_queue.put(("log", "查询超时。任务可能仍在后台处理或已失败。"))
                return
            
            elif task_type in ["t2v", "i2v"]:
                # 原有逻辑保持不变
                submit_id = None
                creation_response = None

                if task_type == "t2v":
                    prompt, ratio, duration_ms = args
                    self.api_queue.put(("log", f"正在提交文生视频任务..."))
                    creation_response = api.create_text_to_video_task(prompt, ratio, duration_ms)
                    submit_id = creation_response.get("data", {}).get("aigc_data", {}).get("submit_id")

                elif task_type == "i2v":
                    prompt, start_path, use_first, end_path, ratio, duration_ms = args
                    self.api_queue.put(("log", "开始处理图生视频任务..."))
                    
                    def _upload_image(image_path, log_name):
                        self.api_queue.put(("log", f"正在读取 {log_name} 图片: {image_path}"))
                        with open(image_path, 'rb') as f: image_bytes = f.read()
                        
                        self.api_queue.put(("log", f"1/4: 为 {log_name} 获取上传Token..."))
                        token_res = api.get_upload_token()
                        if token_res.get("error"): raise Exception(f"获取Token失败: {token_res.get('details')}")
                        
                        self.api_queue.put(("log", f"2/4: 为 {log_name} 申请上传..."))
                        apply_res = api.apply_image_upload(len(image_bytes), token_res["data"])
                        if apply_res.get("error"): raise Exception(f"申请上传失败: {apply_res.get('details')}")
                        
                        upload_info = apply_res["Result"]["UploadAddress"]["StoreInfos"][0]
                        upload_info['UploadHost'] = apply_res["Result"]["UploadAddress"]["UploadHosts"][0]
                        session_key = apply_res["Result"]["UploadAddress"]["SessionKey"]
                        
                        self.api_queue.put(("log", f"3/4: 上传 {log_name} 文件..."))
                        upload_res = api.upload_image_file(image_bytes, upload_info)
                        if upload_res.get("error"): raise Exception(f"上传文件失败: {upload_res.get('error')}")
                        
                        self.api_queue.put(("log", f"4/4: 确认 {log_name} 上传..."))
                        commit_res = api.commit_image_upload(session_key, token_res["data"])
                        if commit_res.get("error"): raise Exception(f"确认上传失败: {commit_res.get('details')}")
                        
                        image_uri = commit_res["Result"]["PluginResult"][0]["ImageUri"]
                        self.api_queue.put(("log", f"{log_name} 上传成功! URI: {image_uri}"))
                        return image_uri

                    start_uri = _upload_image(start_path, "起始帧")
                    end_uri = _upload_image(end_path, "结束帧") if use_first else None
                    
                    self.api_queue.put(("log", "图片上传完成，正在提交视频生成任务..."))
                    creation_response = api.create_image_to_video_task(prompt, start_uri, ratio, "720p", duration_ms, "Video 3.0", use_first, end_uri)
                    submit_id = creation_response.get("data", {}).get("aigc_data", {}).get("submit_id")

                if not submit_id:
                    self.api_queue.put(("log", f"错误：任务提交失败或未能获取submit_id。"))
                    self.api_queue.put(("result", creation_response)); return

                self.api_queue.put(("log", f"任务提交成功！Submit ID: {submit_id}"))
                self.api_queue.put(("log", "开始轮询任务状态... (请耐心等待，可能需几分钟)"))

                max_retries = 30
                for i in range(max_retries):
                    time.sleep(10)
                    self.api_queue.put(("log", f"正在进行第 {i + 1}/{max_retries} 次查询..."))
                    query_response = api.query_video_task_status(submit_id)
                    
                    task_data = query_response.get("data", {}).get(submit_id, {})
                    if not task_data:
                        self.api_queue.put(("log", f"警告：查询响应中未找到ID [{submit_id}] 的数据，继续...")); continue
                    
                    status = task_data.get("status")
                    item_list = task_data.get("item_list", [])
                    self.api_queue.put(("log", f"  查询结果: status={status}, item_list中有{len(item_list)}个项目"))

                    if status != 20:
                        self.api_queue.put(("log", f"任务已结束，最终状态码: {status}"))
                        
                        video_url = None
                        try:
                            if item_list and isinstance(item_list, list) and len(item_list) > 0:
                                item = item_list[0]
                                video_info = item.get("video", {})
                                transcoded_video = video_info.get("transcoded_video", {})
                                origin_video = transcoded_video.get("origin", {})
                                video_url = origin_video.get("video_url")
                        except (AttributeError, IndexError, TypeError) as e:
                            self.api_queue.put(("log", f"解析视频链接时遇到错误: {e}, 将打印原始数据。"))
                        
                        if video_url:
                            self.api_queue.put(("log", "🎉 视频生成成功！"))
                            self.api_queue.put(("log", "============== 视频链接 =============="))
                            self.api_queue.put(("log", video_url))
                            self.api_queue.put(("log", "======================================"))
                        else:
                            self.api_queue.put(("log", f"任务已结束但未能从复杂结构中找到视频链接。"))
                            self.api_queue.put(("result", query_response))
                        return

                self.api_queue.put(("log", "查询超时。任务可能仍在后台处理或已失败。"))
        
        except Exception as e:
            self.api_queue.put(("log", f"发生严重错误: {e}"))
            import traceback
            self.api_queue.put(("log", traceback.format_exc()))
        finally:
            self.api_queue.put(("done", task_type))

    # ========== 账号管理相关方法 ==========
    
    def refresh_account_list(self):
        """刷新账号列表"""
        # 清空现有数据
        for item in self.account_tree.get_children():
            self.account_tree.delete(item)
        
        # 加载所有账号
        accounts = self.account_db.get_all_accounts()
        for acc in accounts:
            status = '✅活跃' if acc['status'] == 'active' else '❌失效'
            self.account_tree.insert('', 'end', values=(
                acc['id'],
                acc['email'],
                acc['password'],
                acc['session_id'][:20] + '...' if len(acc['session_id']) > 20 else acc['session_id'],
                acc['did'],
                acc['create_time'],
                status
            ))
        
        # 更新统计信息
        stats = self.account_db.get_statistics()
        self.account_stats_label.config(
            text=f"统计: 总计 {stats['total']} 个账号 | 可用 {stats['active']} 个 | 失效 {stats['failed']} 个"
        )
    
    def on_account_double_click(self, event):
        """双击账号行复制完整信息到剪贴板"""
        selection = self.account_tree.selection()
        if not selection:
            return
        
        item = self.account_tree.item(selection[0])
        values = item['values']
        
        # 找到完整的账号信息
        account_id = values[0]
        accounts = self.account_db.get_all_accounts()
        account = next((acc for acc in accounts if acc['id'] == account_id), None)
        
        if account:
            # 复制完整信息到剪贴板
            info = f"邮箱: {account['email']}\n密码: {account['password']}\nSession: {account['session_id']}\nDID: {account['did']}"
            self.clipboard_clear()
            self.clipboard_append(info)
            self.log(f"✅ 已复制账号 #{account_id} 的完整信息到剪贴板")
            messagebox.showinfo("复制成功", f"账号 #{account_id} 的完整信息已复制到剪贴板！")
    
    def start_batch_register(self):
        """开始批量注册"""
        count = int(self.batch_count_spinbox.get())
        concurrent = int(self.batch_concurrent_spinbox.get())
        password = self.batch_password_entry.get().strip()
        delay = int(self.batch_delay_spinbox.get())
        
        if not password:
            messagebox.showerror("错误", "请输入密码模板！")
            return
        
        if count <= 0 or concurrent <= 0:
            messagebox.showerror("错误", "注册数量和并发数必须大于0！")
            return
        
        # 确认开始
        result = messagebox.askyesno(
            "确认批量注册",
            f"即将批量注册 {count} 个账号\n"
            f"并发数: {concurrent}\n"
            f"密码: {password}\n"
            f"延迟: {delay} 秒/账号\n\n"
            f"💡 预计耗时: {int(count * delay / concurrent / 60)} 分钟\n\n"
            f"是否继续？"
        )
        
        if not result:
            return
        
        # 禁用开始按钮，启用暂停/停止按钮
        self.batch_start_button.config(state=tk.DISABLED)
        self.batch_pause_button.config(state=tk.NORMAL)
        self.batch_stop_button.config(state=tk.NORMAL)
        
        # 重置进度
        self.batch_progress['maximum'] = count
        self.batch_progress['value'] = 0
        self.batch_status_label.config(text=f"0/{count} 成功:0 失败:0")
        
        # 创建批量注册管理器
        self.batch_manager = BatchRegisterManager(
            count=count,
            concurrent=concurrent,
            password=password,
            database=self.account_db,
            delay=delay,
            progress_callback=self.on_batch_progress
        )
        
        # 在新线程中执行批量注册
        thread = threading.Thread(target=self.batch_manager.batch_register)
        thread.daemon = True
        thread.start()
        
        self.log(f"🚀 开始批量注册 {count} 个账号")
        self.log(f"⚙️ 配置: 并发={concurrent}, 延迟={delay}秒/账号, 密码={password}")
        self.log(f"⏱️ 预计耗时: {int(count * delay / concurrent / 60)} 分钟")
        self.log("=" * 60)
    
    def pause_batch_register(self):
        """暂停/恢复批量注册"""
        if self.batch_manager:
            if self.batch_manager.paused:
                self.batch_manager.resume()
                self.batch_pause_button.config(text="⏸️ 暂停")
                self.log("▶️ 批量注册已恢复")
            else:
                self.batch_manager.pause()
                self.batch_pause_button.config(text="▶️ 恢复")
                self.log("⏸️ 批量注册已暂停")
    
    def stop_batch_register(self):
        """停止批量注册"""
        if self.batch_manager:
            self.batch_manager.stop()
            self.log("⏹️ 批量注册已停止")
    
    def on_batch_progress(self, status, index, data):
        """批量注册进度回调"""
        def update_ui():
            if status == 'log':
                # 实时日志输出
                self.log(f"📝 账号 #{index}: {data}")
                return
            
            if status == 'success':
                # 成功注册一个账号
                self.batch_progress['value'] = self.batch_manager.success_count + self.batch_manager.failed_count
                self.batch_status_label.config(
                    text=f"{self.batch_manager.success_count + self.batch_manager.failed_count}/{self.batch_manager.count} "
                         f"成功:{self.batch_manager.success_count} 失败:{self.batch_manager.failed_count}"
                )
                self.log(f"✅ 账号 #{index} 注册成功: {data['email']}")
                # 刷新列表
                self.refresh_account_list()
                
            elif status == 'failed':
                # 注册失败
                self.batch_progress['value'] = self.batch_manager.success_count + self.batch_manager.failed_count
                self.batch_status_label.config(
                    text=f"{self.batch_manager.success_count + self.batch_manager.failed_count}/{self.batch_manager.count} "
                         f"成功:{self.batch_manager.success_count} 失败:{self.batch_manager.failed_count}"
                )
                self.log(f"❌ 账号 #{index} 注册失败: {data}")
                
            elif status == 'complete':
                # 全部完成
                self.batch_start_button.config(state=tk.NORMAL)
                self.batch_pause_button.config(state=tk.DISABLED, text="⏸️ 暂停")
                self.batch_stop_button.config(state=tk.DISABLED)
                
                self.log("\n" + "=" * 50)
                self.log("🎊 批量注册完成！")
                self.log("=" * 50)
                self.log(f"✅ 成功: {self.batch_manager.success_count} 个")
                self.log(f"❌ 失败: {self.batch_manager.failed_count} 个")
                self.log(f"📊 成功率: {self.batch_manager.success_count / self.batch_manager.count * 100:.1f}%")
                self.log("=" * 50)
                
                # 刷新列表
                self.refresh_account_list()
                
                messagebox.showinfo(
                    "批量注册完成",
                    f"注册完成！\n\n成功: {self.batch_manager.success_count} 个\n失败: {self.batch_manager.failed_count} 个"
                )
        
        # 在主线程中更新UI
        self.after(0, update_ui)
    
    def delete_selected_accounts(self):
        """删除选中的账号"""
        selection = self.account_tree.selection()
        if not selection:
            messagebox.showwarning("提示", "请先选中要删除的账号！")
            return
        
        result = messagebox.askyesno(
            "确认删除",
            f"确定要删除选中的 {len(selection)} 个账号吗？"
        )
        
        if result:
            account_ids = []
            for item in selection:
                values = self.account_tree.item(item)['values']
                account_ids.append(values[0])
            
            self.account_db.delete_selected_accounts(account_ids)
            self.refresh_account_list()
            self.log(f"🗑️ 已删除 {len(account_ids)} 个账号")
    
    def clear_all_accounts(self):
        """清空所有账号"""
        stats = self.account_db.get_statistics()
        if stats['total'] == 0:
            messagebox.showinfo("提示", "账号列表已经是空的！")
            return
        
        result = messagebox.askyesno(
            "确认清空",
            f"确定要清空所有 {stats['total']} 个账号吗？\n此操作不可恢复！"
        )
        
        if result:
            self.account_db.clear_all()
            self.refresh_account_list()
            self.log("🧹 已清空所有账号")
    
    def export_csv(self):
        """导出为CSV文件"""
        stats = self.account_db.get_statistics()
        if stats['total'] == 0:
            messagebox.showinfo("提示", "账号列表为空，无需导出！")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="导出CSV文件",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"capcut_accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if filepath:
            if self.account_db.export_to_csv(filepath):
                self.log(f"📄 成功导出 {stats['total']} 个账号到: {filepath}")
                messagebox.showinfo("导出成功", f"已成功导出 {stats['total']} 个账号到:\n{filepath}")
            else:
                self.log("❌ 导出CSV失败")
                messagebox.showerror("导出失败", "导出CSV文件时发生错误！")
    
    def export_excel(self):
        """导出为Excel文件"""
        stats = self.account_db.get_statistics()
        if stats['total'] == 0:
            messagebox.showinfo("提示", "账号列表为空，无需导出！")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="导出Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"capcut_accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if filepath:
            if self.account_db.export_to_excel(filepath):
                self.log(f"📊 成功导出 {stats['total']} 个账号到: {filepath}")
                messagebox.showinfo("导出成功", f"已成功导出 {stats['total']} 个账号到:\n{filepath}")
            else:
                self.log("❌ 导出Excel失败（可能需要安装 openpyxl 库）")
                messagebox.showwarning("导出失败", "导出Excel失败！\n\n如果未安装 openpyxl 库，可以选择导出为CSV格式。")

    def process_queue(self):
        try:
            while True:
                msg_type, msg_content = self.api_queue.get_nowait()
                if msg_type == "log": self.log(msg_content)
                elif msg_type == "result": self.log(json.dumps(msg_content, indent=4, ensure_ascii=False))
                elif msg_type == "done":
                    if msg_content == "t2i": self.t2i_button.config(state=tk.NORMAL, text="✨ 生成图片 ✨")
                    elif msg_content == "t2v": self.t2v_button.config(state=tk.NORMAL, text="🎬 生成视频 🎬")
                    elif msg_content in ["i2v", "i2v_submit_only", "i2v_query"]: 
                        self.i2v_button.config(state=tk.NORMAL, text="🚀 从图片生成视频 🚀")
                    elif msg_content == "credit": self.credit_button.config(state=tk.NORMAL, text="💰 查询积分")
                    elif msg_content == "credit_receive": self.receive_credit_button.config(state=tk.NORMAL, text="🎁 领取积分")
                    elif msg_content == "config": self.config_button.config(state=tk.NORMAL, text="🔧 获取配置")
                    elif msg_content == "user_info": self.user_info_button.config(state=tk.NORMAL, text="👤 获取用户信息")
                    elif msg_content == "login": self.login_button.config(state=tk.NORMAL, text="🔐 登录获取Session")
                    elif msg_content == "register": self.register_button.config(state=tk.NORMAL, text="📝 注册新账号")
                    elif msg_content == "auto_register": self.auto_register_button.config(state=tk.NORMAL, text="🤖 全自动注册")
        except queue.Empty: pass
        finally: self.after(100, self.process_queue)

if __name__ == "__main__":
    if TK_AVAILABLE:
        app = Application()
        app.mainloop()
    else:
        print("Headless mode: GUI disabled. Core APIs are available for import.")

# --- END OF FILE jimeng003.py (with model selection) ---